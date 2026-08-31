"""The .xlsx 2.4.4 link lane, end to end — extraction, proposal, writer and verification as one.

WHAT WAS BROKEN, and why it had to be fixed in one change rather than four.
tests/test_xlsx_link_credit_gap.py recorded two halves of a single defect:

  1. `proposals.extract_office_links` required a `display=` attribute, while the DETECTOR
     (office_structure.xlsx_structure_checks) resolves a display-less link's label from the cell
     value. openpyxl and most non-Excel generators write hyperlinks with no display attribute, so
     a reviewer routinely saw a 2.4.4 finding with nothing to approve — a live, user-facing gap.
  2. `apply_link_text._apply_xlsx` wrote only the `display` attribute and never touched the cell.
     Since the cell value is what a person reads in the grid, an approved value could clear ACP's
     re-scan while the spreadsheet still showed "click here".

Fixing (1) alone would have made (2) reachable: proposals would finally exist for exactly the
links whose writer could not move the visible text. That is why the owner asked for extraction,
proposal, writer and verification to move together, and why this file exercises all four.

WHAT IS PROVED HERE, precisely. The value of the CELL — the text a spreadsheet reader sees, read
back with openpyxl, a library that had no part in writing it — changes to the approved text;
hyperlink targets, formulas, styles, other cells and the shared-string table survive; the workbook
still opens; a real re-scan no longer reports 2.4.4; and an approved value that is itself vague is
written nowhere and credited nowhere.

WHAT IS NOT PROVED. How Excel and a specific screen reader derive a hyperlink's ACCESSIBLE NAME
from the cell value versus a `display` attribute is a rendered behaviour no static check here can
observe, and nothing in this environment can run Excel. The lane writes the text a sighted reader
sees and keeps the cached attribute in step with it; that is a strictly better position than
before, and it is not the same claim as "verified in Excel with assistive technology".
"""
from __future__ import annotations

import io
import sys
import tempfile
import zipfile
from pathlib import Path

import pytest

ACP = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ACP / "api"))

pytest.importorskip("openpyxl")

import office_structure as _os                                            # noqa: E402
from apply_link_text import apply_link_text                               # noqa: E402
from proposals import extract_office_links, propose_link_texts            # noqa: E402

FILE = "findings.xlsx"
SID = "rv-xlsx-244"

VAGUE_HREF = "https://example.org/q3-accessibility-findings"
GOOD_HREF = "https://example.org/procurement-policy"
VAGUE_TEXT = "click here"
GOOD_TEXT = "procurement accessibility policy"
APPROVED = "Q3 accessibility findings"


# ── fixtures: one workbook per storage shape, because they fail differently ───

def _openpyxl_book(vague: bool = True) -> bytes:
    """What a generator actually produces: hyperlinks with NO display attribute, an unrelated
    cell holding the same text, a formula, and a styled header."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"
    ws["A1"] = "Report"
    ws["A1"].font = Font(bold=True)
    ws["B1"] = "Link"
    ws["A2"] = "Q3 summary"
    ws["B2"] = VAGUE_TEXT if vague else APPROVED
    ws["B2"].hyperlink = VAGUE_HREF
    ws["B3"] = GOOD_TEXT
    ws["B3"].hyperlink = GOOD_HREF
    ws["C2"] = "=LEN(A2)"
    # An unrelated cell carrying the SAME text as the vague label. In a shared-string workbook
    # this is the cell a naive in-place edit would corrupt; here it also catches a writer that
    # rewrites by text-match instead of by cell reference.
    ws["A5"] = VAGUE_TEXT
    with tempfile.TemporaryDirectory() as d:
        p = Path(d) / FILE
        wb.save(p)
        return p.read_bytes()


_SS = ('<?xml version="1.0"?><sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
       'count="4" uniqueCount="3">'
       '<si><t>click here</t></si>'          # index 0 — SHARED by B2 (the link) and A5 (not a link)
       '<si><t>Q3 summary</t></si>'
       '<si><t>Notes</t></si>'
       '</sst>')

_SHEET = ('<?xml version="1.0"?><worksheet '
          'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
          'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheetData>'
          '<row r="2"><c r="A2" t="s"><v>1</v></c><c r="B2" s="4" t="s"><v>0</v></c></row>'
          '<row r="5"><c r="A5" t="s"><v>0</v></c><c r="B5" t="s"><v>2</v></c></row>'
          '</sheetData>'
          '<hyperlinks><hyperlink ref="B2" r:id="rId1"/></hyperlinks></worksheet>')

_RELS = ('<?xml version="1.0"?><Relationships '
         'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
         f'<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/'
         f'relationships/hyperlink" Target="{VAGUE_HREF}" TargetMode="External"/></Relationships>')


def _shared_string_book() -> bytes:
    """What EXCEL writes: cell values live in xl/sharedStrings.xml, and one <si> is referenced by
    two cells. openpyxl emits inline strings instead, so this shape has to be built by hand —
    and it is the one where an in-place edit silently corrupts an unrelated cell."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("xl/worksheets/sheet1.xml", _SHEET)
        z.writestr("xl/worksheets/_rels/sheet1.xml.rels", _RELS)
        z.writestr("xl/sharedStrings.xml", _SS)
    return buf.getvalue()


def _spill(data: bytes, name: str = FILE) -> str:
    f = tempfile.NamedTemporaryFile(suffix=Path(name).suffix, delete=False)
    f.write(data)
    f.close()
    return f.name


def _part(data: bytes, name: str) -> str:
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        return z.read(name).decode("utf-8")


def _cell_values(data: bytes) -> dict[str, str]:
    """Resolve every cell's text the way the DETECTOR does, through the shared-string table."""
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        names = z.namelist()
        ss = (_os._read(z, "xl/sharedStrings.xml") or "") if "xl/sharedStrings.xml" in names else ""
        sheet = _os._read(z, "xl/worksheets/sheet1.xml") or ""
    return _os._xlsx_cell_text_values(sheet, _os._xlsx_shared_strings(ss) if ss else [])


def _scs(path: str) -> set[str]:
    return {(f.get("wcag") or "").split()[0] for f in _os.xlsx_structure_checks(Path(path))}


class _Blob:
    """Stands in for Azure Blob, and it is the only thing this module patches. It stores bytes
    verbatim and decides nothing about whether the criterion cleared."""

    def __init__(self, data: bytes):
        self.data, self.uploads = data, []

    def download_remediated(self, owner, sid, f):
        return self.data

    def upload_remediated(self, owner, sid, f, data, mime):
        self.data = data
        self.uploads.append((f, mime))
        return "http://b/2"


@pytest.fixture(scope="module")
def book() -> bytes:
    return _openpyxl_book()


# ── 1. the finding, and the proposal that was missing ────────────────────────

def test_the_detector_reports_2_4_4_on_a_display_less_link(book):
    assert "2.4.4" in _scs(_spill(book))


def test_extraction_now_reads_a_display_less_link(book):
    """Gap 1. The label of a link with no display attribute is the value of its cell, which is
    how the detector has always read it. Extraction agreeing is what gives the finding a fix."""
    links = dict((h, t) for t, h in extract_office_links(_spill(book), "xlsx"))
    assert links.get(VAGUE_HREF) == VAGUE_TEXT, (
        "the vague link is invisible to extraction, so its finding reaches a reviewer with "
        "nothing to approve")
    assert links.get(GOOD_HREF) == GOOD_TEXT


def test_a_reviewer_is_offered_something_to_approve(book):
    props = [p for p in propose_link_texts(_spill(book), "xlsx", ai_enabled=False)
             if p["locator"] == VAGUE_HREF]
    assert props, "a 2.4.4 finding with no proposal is a finding a reviewer cannot act on"
    assert props[0]["before"] == VAGUE_TEXT
    assert props[0]["proposed_value"].strip()
    assert props[0]["sc"] == "2.4.4"


def test_the_descriptive_link_is_not_proposed_for(book):
    """The control. A proposer that offered work on every link would satisfy the test above."""
    assert GOOD_HREF not in {p["locator"]
                             for p in propose_link_texts(_spill(book), "xlsx", ai_enabled=False)}


# ── 2. the write: the text a reader sees ─────────────────────────────────────

@pytest.fixture(scope="module")
def written(book) -> bytes:
    out, applied, unresolved = apply_link_text(book, "xlsx", {VAGUE_HREF: APPROVED})
    assert unresolved == [], unresolved
    assert [a["locator"] for a in applied] == [VAGUE_HREF]
    assert applied[0]["before"] == VAGUE_TEXT and applied[0]["after"] == APPROVED
    return out


def test_the_cell_a_reader_sees_now_carries_the_approved_text(written):
    """Gap 2, and the whole point. Previously only `display` moved, so the grid still read
    "click here" while ACP recorded the criterion as cleared."""
    assert _cell_values(written)["B2"] == APPROVED


def test_openpyxl_reads_the_new_value_back(written):
    """Independent confirmation, through a library that had no part in writing the file."""
    from openpyxl import load_workbook
    ws = load_workbook(_spill(written))["Findings"]
    assert ws["B2"].value == APPROVED
    assert ws["B2"].hyperlink.target == VAGUE_HREF, "the destination moved"


def test_the_unrelated_cell_holding_the_same_text_is_untouched(written):
    """A5 says "click here" too and is not a link. A writer matching on text rather than on the
    hyperlink's own cell reference would rewrite it."""
    assert _cell_values(written)["A5"] == VAGUE_TEXT


def test_formulas_styles_and_other_content_survive(written):
    from openpyxl import load_workbook
    ws = load_workbook(_spill(written))["Findings"]
    assert ws["C2"].value == "=LEN(A2)", "the formula was clobbered"
    assert ws["A1"].font.bold is True, "cell formatting was dropped"
    assert ws["A2"].value == "Q3 summary"
    assert ws["B3"].value == GOOD_TEXT and ws["B3"].hyperlink.target == GOOD_HREF


def test_the_workbook_still_opens(written):
    assert zipfile.ZipFile(io.BytesIO(written)).testzip() is None
    from openpyxl import load_workbook
    assert load_workbook(_spill(written)).sheetnames == ["Findings"]


def test_a_rescan_of_the_written_file_no_longer_reports_2_4_4(written):
    assert "2.4.4" not in _scs(_spill(written))


# ── 3. the shared-string trap ────────────────────────────────────────────────

def test_a_shared_string_is_never_edited_in_place():
    """THE correctness trap of this format. A t="s" cell holds an INDEX into a table any number
    of other cells may share, so editing the <si> would rewrite every one of them. B2 (the link)
    and A5 (unrelated) both point at index 0 here."""
    out, applied, unresolved = apply_link_text(_shared_string_book(), "xlsx",
                                               {VAGUE_HREF: APPROVED})
    assert unresolved == [] and len(applied) == 1
    values = _cell_values(out)
    assert values["B2"] == APPROVED
    assert values["A5"] == VAGUE_TEXT, (
        "an unrelated cell sharing the string was rewritten — the <si> was edited in place")
    assert values["A2"] == "Q3 summary" and values["B5"] == "Notes"


def test_the_original_entry_survives_and_a_new_one_is_appended():
    out, _, _ = apply_link_text(_shared_string_book(), "xlsx", {VAGUE_HREF: APPROVED})
    ss = _part(out, "xl/sharedStrings.xml")
    assert "<si><t>click here</t></si>" in ss, "the entry other cells depend on was removed"
    assert f"<si><t>{APPROVED}</t></si>" in ss
    assert _os._xlsx_shared_strings(ss) == ["click here", "Q3 summary", "Notes", APPROVED]


def test_the_shared_string_counts_stay_consistent():
    """`uniqueCount` counts <si> entries and grows by one. `count` counts string-cell REFERENCES
    and must not move: repointing a cell relocates a reference rather than adding one."""
    out, _, _ = apply_link_text(_shared_string_book(), "xlsx", {VAGUE_HREF: APPROVED})
    ss = _part(out, "xl/sharedStrings.xml")
    assert 'uniqueCount="4"' in ss
    assert 'count="4"' in ss                       # the original total, unchanged


def test_the_cells_style_survives_the_repoint():
    out, _, _ = apply_link_text(_shared_string_book(), "xlsx", {VAGUE_HREF: APPROVED})
    sheet = _part(out, "xl/worksheets/sheet1.xml")
    assert 's="4"' in sheet, "the cell's style index was lost when its value was repointed"


# ── 4. what the writer must REFUSE ───────────────────────────────────────────

def _formula_label_book() -> bytes:
    sheet = ('<?xml version="1.0"?><worksheet '
             'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
             'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
             '<sheetData><row r="2">'
             '<c r="B2" t="str"><f>A1&amp;" here"</f><v>click here</v></c>'
             '</row></sheetData>'
             '<hyperlinks><hyperlink ref="B2" r:id="rId1"/></hyperlinks></worksheet>')
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("xl/worksheets/sheet1.xml", sheet)
        z.writestr("xl/worksheets/_rels/sheet1.xml.rels", _RELS)
    return buf.getvalue()


def test_a_formula_label_is_reported_never_overwritten():
    """A computed label cannot be replaced with literal text: overwriting the cached <v> is undone
    on the next recalculation, and replacing <f> destroys the formula. Refusing and SAYING SO is
    the honest outcome — the reviewer learns their approval did not land, and the row stays
    unapplied so the file cannot certify off it."""
    data = _formula_label_book()
    out, applied, unresolved = apply_link_text(data, "xlsx", {VAGUE_HREF: APPROVED})
    assert applied == [], "a formula cell was rewritten as literal text"
    assert out == data, "the document changed despite nothing being applied"
    assert len(unresolved) == 1
    assert "formula" in unresolved[0], unresolved[0]
    assert VAGUE_HREF in unresolved[0], "the reviewer is not told WHICH approval did not land"


def test_a_refused_link_does_not_get_its_display_attribute_written_instead():
    """The tempting half-fix, and the exact divergence this change exists to end: writing the
    cached attribute while the cell a reader sees keeps the old text would clear ACP's own
    re-scan on a document nobody fixed."""
    sheet = _formula_label_book()
    out, _, _ = apply_link_text(sheet, "xlsx", {VAGUE_HREF: APPROVED})
    assert f'display="{APPROVED}"' not in _part(out, "xl/worksheets/sheet1.xml")


def test_an_already_descriptive_workbook_is_left_alone():
    book = _openpyxl_book(vague=False)
    assert "2.4.4" not in _scs(_spill(book))
    assert [p for p in propose_link_texts(_spill(book), "xlsx", ai_enabled=False)
            if p["locator"] == VAGUE_HREF] == []
    out, applied, _ = apply_link_text(book, "xlsx", {})
    assert applied == [] and out == book


# ── 5. the display= shape still works ────────────────────────────────────────

def test_a_link_authored_with_a_display_attribute_still_has_it_rewritten():
    """Excel writes `display` as a cache of the cell text, and a sheet may carry the attribute
    with no cell at all. Rewriting the cell must not have cost that path its write — an earlier
    draft of this change refused every display-only link, because "no cell at B2" was treated as
    a failure rather than as "there is nothing here that needs one"."""
    sheet = ('<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
             '<hyperlinks><hyperlink ref="A1" r:id="rId1" display="click here"/></hyperlinks>'
             '</worksheet>')
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("xl/worksheets/sheet1.xml", sheet)
        z.writestr("xl/worksheets/_rels/sheet1.xml.rels", _RELS)
    out, applied, unresolved = apply_link_text(buf.getvalue(), "xlsx", {VAGUE_HREF: APPROVED})
    assert unresolved == [] and len(applied) == 1
    assert f'display="{APPROVED}"' in _part(out, "xl/worksheets/sheet1.xml")


def test_the_whole_lane_runs_through_the_handler_with_the_rescan_unpatched(monkeypatch, book):
    """THE bar for REMEDIATION-VERIFIED, and the reason this section exists.

    Everything above drives `apply_link_text` and `xlsx_structure_checks` directly. That proves
    the writer and the detector, and it is NOT what the level claims — which is that an approved
    value goes through the PRODUCTION lane and the saved document is re-scanned by the real path
    (handlers._apply_approved_values -> proposals.verify_residual_scs -> scanner.analyse_and_assess).
    A first draft of this file stopped at the direct calls and still added the entry; the guard in
    tests/test_capability_levels.py did not catch it, because it only refused a test that PATCHES
    the re-scan and this one simply never reached it. Both are fixed here.

    Nothing is patched but the blob store, which decides nothing.
    """
    import core
    import handlers

    from proposals import propose_link_texts
    props = [p for p in propose_link_texts(_spill(book), "xlsx", ai_enabled=False)
             if p["locator"] == VAGUE_HREF]
    assert props, "nothing to approve — the earlier proposal tests should have caught this"

    import store as store_mod
    monkeypatch.setattr(store_mod, "_SQLITE_PATH", Path(tempfile.mkdtemp()) / "rv.db")
    store = store_mod.Store()
    store.init_scan_run(SID, "drive", 1, "2026-08-30T00:00:00Z", "rubric", "hash")
    store.save_file_result(SID, {
        "file": FILE, "engine": "office", "status": "pass", "score": 60, "compliant": 0,
        "skipped_rules": 0, "drive_file_id": "drv1",
        "issues": [{"ruleId": "XLSX_LINK_PURPOSE_VAGUE", "wcag": "2.4.4 Link Purpose",
                    "severity": "MODERATE"}],
    }, "2026-08-30T00:00:00Z")
    store.record_remediation(SID, FILE, drive_write_url="http://d/1", blob_url="http://b/1")
    item_id = store.enqueue_proposals(SID, FILE, "2.4.4", [
        {k: p[k] for k in ("locator", "before", "proposed_value", "rationale", "source")}
        for p in props], rule_name="Link Purpose (In Context)")
    store.update_hitl_item(item_id, "approved", None, None)
    store.approve_proposal_values(item_id, [APPROVED])

    blob = _Blob(book)
    monkeypatch.setattr(core, "store", store)
    monkeypatch.setitem(sys.modules, "blob", blob)
    handlers._apply_approved_values({"scan_id": SID, "file": FILE}, {})

    assert blob.uploads, "the corrected copy was never stored"
    assert store.count_unapplied_approved_values(SID, FILE) == 0, (
        "the approved value was written but never credited")
    from openpyxl import load_workbook
    assert load_workbook(_spill(blob.data))["Findings"]["B2"].value == APPROVED


def test_the_handler_withholds_credit_when_the_approved_text_is_itself_vague(monkeypatch, book):
    """The control that distinguishes a real re-scan from a supplied one. The write succeeds, the
    re-scan still reports 2.4.4, and nothing may be credited or published."""
    import core
    import handlers
    import store as store_mod

    from proposals import propose_link_texts
    props = [p for p in propose_link_texts(_spill(book), "xlsx", ai_enabled=False)
             if p["locator"] == VAGUE_HREF]
    monkeypatch.setattr(store_mod, "_SQLITE_PATH", Path(tempfile.mkdtemp()) / "rv2.db")
    store = store_mod.Store()
    store.init_scan_run(SID, "drive", 1, "2026-08-30T00:00:00Z", "rubric", "hash")
    store.save_file_result(SID, {
        "file": FILE, "engine": "office", "status": "pass", "score": 60, "compliant": 0,
        "skipped_rules": 0, "drive_file_id": "drv1",
        "issues": [{"ruleId": "XLSX_LINK_PURPOSE_VAGUE", "wcag": "2.4.4 Link Purpose",
                    "severity": "MODERATE"}],
    }, "2026-08-30T00:00:00Z")
    store.record_remediation(SID, FILE, drive_write_url="http://d/1", blob_url="http://b/1")
    item_id = store.enqueue_proposals(SID, FILE, "2.4.4", [
        {k: p[k] for k in ("locator", "before", "proposed_value", "rationale", "source")}
        for p in props], rule_name="Link Purpose (In Context)")
    store.update_hitl_item(item_id, "approved", None, None)
    store.approve_proposal_values(item_id, ["read more"])          # still vague

    blob = _Blob(book)
    monkeypatch.setattr(core, "store", store)
    monkeypatch.setitem(sys.modules, "blob", blob)
    handlers._apply_approved_values({"scan_id": SID, "file": FILE}, {})

    assert store.count_unapplied_approved_values(SID, FILE) == 1, (
        "a value that did not clear the criterion was credited anyway")
    assert store.mark_file_compliant_if_reviewed(SID, FILE) is False
    assert not blob.uploads, "an uncleared write was published as the corrected copy"


def test_display_and_cell_are_kept_in_step_when_both_exist():
    """When a sheet carries both, they must not be allowed to disagree — a stale `display` next
    to a fixed cell is the same divergence in the other direction."""
    sheet = ('<?xml version="1.0"?><worksheet '
             'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
             'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
             '<sheetData><row r="2"><c r="B2" t="inlineStr"><is><t>click here</t></is></c></row>'
             '</sheetData>'
             '<hyperlinks><hyperlink ref="B2" r:id="rId1" display="click here"/></hyperlinks>'
             '</worksheet>')
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("xl/worksheets/sheet1.xml", sheet)
        z.writestr("xl/worksheets/_rels/sheet1.xml.rels", _RELS)
    out, applied, _ = apply_link_text(buf.getvalue(), "xlsx", {VAGUE_HREF: APPROVED})
    xml = _part(out, "xl/worksheets/sheet1.xml")
    assert len(applied) == 1
    assert f'display="{APPROVED}"' in xml
    assert f"<t>{APPROVED}</t>" in xml
    assert "click here" not in xml
