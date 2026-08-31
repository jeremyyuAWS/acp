"""A location is only useful if it reaches the person doing the review.

WHY THIS FILE EXISTS, and it is a correction. tests/test_finding_location.py proves the DETECTORS
now emit a location. From that I claimed the rest of the chain already worked, on the strength of
reading two things that were each true on their own:

  * store._loc() persists `location` (or `locator`) into issue_records — true;
  * EvidenceCard.jsx renders `card.location` as a 📍 chip — also true.

The join between them was not checked, and it did not hold. `card.location` is built by
reviewCard.locationLabel(item), which reads `item.pages` — a comma-separated list of INTEGERS
taken from issue_records.page. The Office detectors set `location` and leave `page` null, because
a worksheet has no page number. So the queue row carried page=None, pages=None, no location at
all, and the chip rendered nothing for every Office finding.

Measured, which is how it surfaced:

    1. detector finding    location = 'Sheet “Q3 Findings” cell B2'   page = None
    2. issue_records row   ('2.4.4 Link Purpose (In Context)', None, 'Sheet “Q3 Findings” cell B2')
    3. hitl item created   {'rule_id': '2.4.4', 'page': None, 'pages': None}     <- lost here

The location reached the database, the report and the file drawer, and stopped one row short of
the review card. This file asserts the whole chain rather than its ends.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "api"))

pytest.importorskip("openpyxl")

import office_structure as osx  # noqa: E402

SID = "loc-chain"
FILE = "book.xlsx"


@pytest.fixture()
def store(monkeypatch):
    import store as store_mod
    monkeypatch.setattr(store_mod, "_SQLITE_PATH", Path(tempfile.mkdtemp()) / "loc.db")
    return store_mod.Store()


def _vague_book(tmp_path: Path) -> Path:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Q3 Findings"
    ws["B2"] = "click here"
    ws["B2"].hyperlink = "https://example.org/report"
    p = tmp_path / FILE
    wb.save(p)
    return p


def _seed_from_real_detector(store, tmp_path) -> dict:
    """Persist a finding produced by the ACTUAL detector, not a hand-written dict — so a change
    in the shape the detector emits breaks this rather than sliding past a fixture."""
    findings = osx.xlsx_structure_checks(_vague_book(tmp_path))
    finding = next(f for f in findings if f.get("ruleId") == "XLSX_LINK_PURPOSE_VAGUE")
    assert finding.get("location"), "the detector stopped emitting a location"
    store.init_scan_run(SID, "drive", 1, "2026-08-31T00:00:00Z", "rubric", "hash")
    store.save_file_result(SID, {
        "file": FILE, "engine": "office", "status": "pass", "score": 60, "compliant": 0,
        "skipped_rules": 0, "drive_file_id": "d1", "issues": [finding],
    }, "2026-08-31T00:00:00Z")
    return finding


def test_the_location_survives_into_the_stored_finding(store, tmp_path):
    finding = _seed_from_real_detector(store, tmp_path)
    rows = store.list_issues(SID) if hasattr(store, "list_issues") else None
    if rows is None:                       # fall back to the column the report reads
        import sqlite3
        import store as store_mod
        con = sqlite3.connect(store_mod._SQLITE_PATH)
        rows = con.execute("SELECT location FROM issue_records WHERE scan_id=? AND file=?",
                           (SID, FILE)).fetchall()
        assert rows and rows[0][0] == finding["location"]
        return
    assert any(r.get("location") == finding["location"] for r in rows)


def test_the_location_reaches_the_review_queue_row(store, tmp_path):
    """THE step that was missing. The queue row is what the review card is built from, and it
    carried only the integer page — null for every Office format."""
    finding = _seed_from_real_detector(store, tmp_path)
    created = store.queue_hitl_items(SID)
    assert created, "nothing was queued, so this test cannot see the row it is about"
    row = created[0]
    assert row.get("location") == finding["location"], (
        f"the queue row lost the location: {row.get('location')!r}")
    assert row.get("pages") in (None, ""), (
        "this fixture is supposed to have NO page number — if it now has one, the fallback "
        "under test is no longer the thing being exercised")


def test_the_listing_the_api_serves_carries_it_too(store, tmp_path):
    """The card reads the LISTING, not the insert's return value."""
    finding = _seed_from_real_detector(store, tmp_path)
    store.queue_hitl_items(SID)
    items = store.list_hitl_queue(scan_id=SID)
    assert items, "the listing is empty"
    assert any(i.get("location") == finding["location"] for i in items), (
        f"no listed item carries the location; got {[i.get('location') for i in items]}")


def test_a_finding_with_no_location_stays_null_rather_than_inventing_one(store, tmp_path):
    """The rule the whole change follows: no position is reported as none, never as a guess."""
    store.init_scan_run(SID, "drive", 1, "2026-08-31T00:00:00Z", "rubric", "hash")
    store.save_file_result(SID, {
        "file": FILE, "engine": "office", "status": "pass", "score": 60, "compliant": 0,
        "skipped_rules": 0, "drive_file_id": "d1",
        "issues": [{"ruleId": "XLSX_IMAGE_NO_ALT", "wcag": "1.1.1 Non-text Content",
                    "severity": "SERIOUS"}],
    }, "2026-08-31T00:00:00Z")
    created = store.queue_hitl_items(SID)
    for row in created:
        assert not row.get("location"), (
            f"a finding that reported no position was given one: {row.get('location')!r}")
