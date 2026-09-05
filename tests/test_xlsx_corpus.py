"""The labelled .xlsx ground-truth corpus — and the check that its labels are earned.

`scripts/gen_fixture_coverage.py` reported 15 of 62 pairs covered: .docx complete, xlsx / pptx /
pdf at zero. This is the start of xlsx, and it is deliberately partial — four pairs, not fifteen.

THE LIMIT IS VERIFICATION, NOT EFFORT, and that distinction is the whole reason this file matters
more than the generator. Coverage is counted from DECLARATIONS, so a fixture that declares
"1.4.3: FAIL" while seeding something no detector catches would raise the coverage number without
raising the coverage. These tests refuse that trade: every violation fixture is driven through
the real first-party detector and asserted to fire, and every adversarial fixture is asserted to
stay silent. A declaration in the manifest is therefore a measured fact.

The remaining eleven .xlsx pairs are absent because their detection runs through the .NET Office
analyser, or (3.1.2) through langdetect — neither installable in every environment, and a label
nobody can confirm is exactly what this file exists to prevent.

WHY THE ADVERSARIAL FIXTURES ARE HALF THE CORPUS. A corpus of obvious violations measures almost
nothing: anything finds white-on-white text. What separates a detector from a regex is the
near-miss — a LONE default 'Sheet1' tab, which Excel gives every new workbook and which must not
be flagged; an iconSet, which pairs colour WITH a shape and is therefore not colour-only. Each of
those names the specific wrong answer it is built to catch.
"""
from __future__ import annotations

import importlib.util
import sys
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "api"))
sys.path.insert(0, str(ROOT / "scripts"))

import office_structure as osx  # noqa: E402
import ocr as _ocr  # noqa: E402
import pii as _pii  # noqa: E402
import textchecks as _tc  # noqa: E402

_spec = importlib.util.spec_from_file_location(
    "gen_xlsx_corpus", ROOT / "scripts" / "gen_xlsx_corpus.py")
gen = importlib.util.module_from_spec(_spec)
sys.modules["gen_xlsx_corpus"] = gen
_spec.loader.exec_module(gen)

import corpus_expectations as ce  # noqa: E402
from engines import NO_OFFICE, OFFICE_OK  # noqa: E402


@pytest.fixture(scope="module")
def corpus(tmp_path_factory):
    """Build every fixture once. Generated, not committed: a generated corpus cannot drift from
    the generator that documents what each fixture is FOR, and it keeps binaries whose provenance
    a reviewer would have to take on trust out of the repo."""
    out = tmp_path_factory.mktemp("xlsx-corpus")
    manifest, problems = gen.build_all(out / "docs")
    assert not problems, f"fixtures declare verdicts the engine cannot emit: {problems}"
    return out, {row["name"]: row for row in manifest}


def _ocr_wcags(path: Path, ext: str) -> set[str]:
    """Criteria the OCR lane reports — 1.4.5, read out of the document's PIXELS rather than its
    structure. A real scan runs this pass alongside the structural one, so `_wcags` is their
    union; checking only the structural lane would make every 1.4.5 fixture invisible and every
    "this fixture is single-criterion" assertion below weaker than it reads.

    Empty when tesseract is unavailable, which is why the 1.4.5 rows below cannot fail a bare
    checkout. Both CI pipelines run scripts/install_tesseract.sh, so that is a fallback and not
    the normal state — see test_ocr_is_present_in_ci."""
    return {(f.get("wcag") or "").split()[0] for f in _ocr.images_of_text(path, ext)
            if f.get("wcag")}


def _text_wcags(path: Path, ext: str) -> set[str]:
    """Criteria the TEXT lane reports — 1.3.3 and 3.1.2, decided by the document's PROSE rather
    than its structure or its pixels. These two calls are what scanner.py makes (scanner.py:3483):
    extract the text, read the document's own language marks, then judge.

    `language_marked_spans` is passed rather than omitted because 3.1.2 asks whether a foreign
    passage's language is IDENTIFIED, and dropping the marks would make a correctly-marked
    document indistinguishable from an unmarked one — the detector would fire on both and no
    control could ever be clean."""
    text = _pii.extract_text(path) or ""
    return {(f.get("wcag") or "").split()[0]
            for f in _tc.content_findings(text, osx.language_marked_spans(path, ext))
            if f.get("wcag")}


def _wcags(path: Path) -> set[str]:
    """Every criterion a real scan of this file reports, across BOTH lanes: the first-party xlsx
    structure checks (pure Python, no external engine) and the OCR pass over its embedded images.
    The union is what makes the single-criterion assertions below mean anything."""
    structural = {(f.get("wcag") or "").split()[0] for f in osx.checks_for(path, ".xlsx")
                  if f.get("wcag")}
    return structural | _ocr_wcags(path, ".xlsx") | _text_wcags(path, ".xlsx")


# ── the labels are legal for their lane ──────────────────────────────────────────

def test_every_declaration_is_a_verdict_the_engine_can_emit(corpus):
    """The generator enforces this at build time; asserting it here too means a fixture cannot be
    added with an impossible label by someone who never runs the generator directly."""
    _out, rows = corpus
    for name, row in rows.items():
        for sc, verdict in row["expect"].items():
            allowed = ce.possible_verdicts(sc, "xlsx")
            assert verdict in allowed, (
                f"{name} expects {sc}={verdict}, but ({sc}, xlsx) can only emit "
                f"{sorted(allowed)}")


def test_no_review_lane_fixture_claims_pass(corpus):
    """The specific trap corpus_expectations exists for. 1.4.1, 2.4.4 and 2.4.6 are review-lane on
    .xlsx: a CLEAN file there resolves to REVIEW and never to PASS, so an adversarial fixture that
    expected PASS would report a false failure forever and read as a product defect."""
    _out, rows = corpus
    for name, row in rows.items():
        for sc, verdict in row["expect"].items():
            if not ce.can_ever_pass(sc, "xlsx"):
                assert verdict != "PASS", (
                    f"{name} expects PASS on {sc}, which .xlsx cannot certify (ADR 0016)")


# ── the labels are EARNED: the detector actually fires ───────────────────────────


# Criteria that ONLY the OCR lane can report — read out of a page's PIXELS, so unreachable
# without tesseract however correct the fixture is. Kept as an explicit set rather than inferred:
# inferring it would mean running the lane to find out, which is the thing being gated.
_OCR_ONLY_SC = {"1.4.5", "1.4.9"}


@pytest.mark.parametrize("name,sc", [
    ("contrast-fail", "1.4.3"),
    ("link-vague", "2.4.4"),
    ("sheet-tabs-default", "2.4.6"),
    ("colour-scale-only", "1.4.1"),
    ("image-no-alt", "1.1.1"),
    ("image-of-text", "1.4.5"),
    ("sensory-instruction", "1.3.3"),
    ("language-parts", "3.1.2"),
    ("shape-faint-outline", "1.4.11"),
    ("form-control", "4.1.2"),
    ("form-control", "2.1.2"),
])
def test_each_violation_fixture_is_actually_detected(corpus, name, sc):
    """Without this, a declaration is a hope. Coverage is counted from declarations, so an
    undetected fixture would raise the number #1009 reports without raising what it measures."""
    # THE SKIP THE DOCSTRING ALREADY PROMISED. `_ocr_wcags` says the 1.4.5 assertions "skip
    # rather than fail on a bare checkout"; they did not — there was no guard here, so a
    # developer without tesseract got three hard FAILURES from a complete, correct checkout.
    # That is worse than noise: a suite that is red for an environmental reason trains everyone
    # to read red as "probably the usual three", which is exactly how a real regression gets
    # waved through.
    #
    # Skipping loses no coverage, because losing it in CI is caught separately and loudly:
    # test_ocr_is_present_in_ci asserts _ocr.is_available() whenever CI/TF_BUILD is
    # set, so a pipeline that stopped installing tesseract fails there rather than quietly
    # skipping here. Both halves are needed — this one keeps a bare checkout honest, that one
    # keeps CI honest.
    if sc in _OCR_ONLY_SC and not _ocr.is_available():
        pytest.skip(f"{sc} is only reachable through the OCR lane and tesseract is unavailable")
    out, _rows = corpus
    got = _wcags(out / "docs" / f"{name}.xlsx")
    assert sc in got, (
        f"{name} seeds a {sc} violation that no first-party detector catches — the manifest "
        f"would claim coverage the corpus does not have. Detected: {sorted(got) or 'nothing'}")


@pytest.mark.parametrize("name,sc", [
    ("link-descriptive-ok", "2.4.4"),
    ("sheet-tab-single-ok", "2.4.6"),
    ("sheet-tabs-named-ok", "2.4.6"),
    ("colour-icon-set-ok", "1.4.1"),
    ("contrast-ok", "1.4.3"),
    ("no-image-ok", "1.1.1"),
    ("image-of-text-logo-ok", "1.4.5"),
    ("sensory-instruction-ok", "1.3.3"),
    ("language-parts-ok", "3.1.2"),
    ("shape-strong-outline-ok", "1.4.11"),
    ("no-controls-ok", "4.1.2"),
    ("no-controls-ok", "2.1.2"),
])
def test_each_adversarial_fixture_stays_silent(corpus, name, sc):
    """The half that separates a detector from a regex. A false positive is cheap to ship and
    expensive to trust: it is what teaches a reviewer to stop reading the findings."""
    out, _rows = corpus
    got = _wcags(out / "docs" / f"{name}.xlsx")
    assert sc not in got, (
        f"{name} is a false positive on {sc} — it is the case the detector is supposed to let "
        f"through. Detected: {sorted(got)}")


def test_one_control_answers_for_both_its_criteria(corpus):
    """An embedded control is evidence for the accessible-name question AND the keyboard-trap
    question, and the detector reports both — so one fixture legitimately covers two pairs.
    Asserted together because a change that dropped either would still pass the other's
    parametrised case above and look fine."""
    out, _rows = corpus
    got = _wcags(out / "docs" / "form-control.xlsx")
    assert {"4.1.2", "2.1.2"} <= got, f"the control fired only {sorted(got)}"


def test_the_faint_outline_is_measured_not_assumed(corpus):
    """Both shape fixtures carry the SAME shape; only the outline colour differs. If the detector
    ever started reporting on the presence of a shape rather than on its measured ratio, the
    adversarial case would fire too — which is the whole distinction 1.4.11 turns on."""
    out, _rows = corpus
    assert "1.4.11" in _wcags(out / "docs" / "shape-faint-outline.xlsx")
    assert "1.4.11" not in _wcags(out / "docs" / "shape-strong-outline-ok.xlsx")


def test_the_lone_default_tab_is_the_edge_case_it_claims_to_be(corpus):
    """Named separately because it is the one most likely to be 'fixed' into a bug. Excel titles
    every new workbook's first sheet 'Sheet1'; flagging one would fire on most files in an
    estate, so the detector requires two before it says anything. Three tabs fire; one does not."""
    out, _rows = corpus
    assert "2.4.6" in _wcags(out / "docs" / "sheet-tabs-default.xlsx")
    assert "2.4.6" not in _wcags(out / "docs" / "sheet-tab-single-ok.xlsx")


# ── the corpus and the coverage report agree ─────────────────────────────────────

def test_the_coverage_report_counts_this_corpus(corpus):
    """The generator, the report and the ratchet's baseline are one fact in three places; this is
    what stops them drifting."""
    import gen_fixture_coverage as gfc
    cov = gfc.coverage()
    assert cov["xlsx"]["has_generator"] is True, (
        "gen_fixture_coverage does not know about the xlsx corpus — add it to GENERATORS")
    declared = set(gen.DECLARED) | set(gen.DECLARED_ENGINE)
    assert sorted(cov["xlsx"]["covered"]) == sorted(declared)
    assert gfc.BASELINE["xlsx"] == len(declared), (
        f"BASELINE['xlsx'] is {gfc.BASELINE['xlsx']} but the corpus declares "
        f"{len(declared)} — the ratchet would have slack in it")
    assert sorted(cov["xlsx"]["engine_only"]) == sorted(gen.DECLARED_ENGINE), (
        "the report's engine-only split disagrees with the generator — the headline number "
        "would then imply a guarantee two of these pairs do not have")


def test_the_declared_set_matches_what_the_fixtures_actually_declare(corpus):
    """DECLARED is written down for the report to read; this keeps it honest against the
    fixtures themselves."""
    _out, rows = corpus
    from_fixtures = {sc for row in rows.values() for sc in row["expect"]}
    assert from_fixtures == set(gen.DECLARED) | set(gen.DECLARED_ENGINE)


# ── the engine-verified pairs: structure here, detection in CI ───────────────────

def _core_xml(path: Path) -> str:
    import zipfile
    with zipfile.ZipFile(path) as z:
        return z.read("docProps/core.xml").decode("utf-8")


def _prop(xml: str, tag: str) -> str | None:
    """A core-property value, or None when the element is absent.

    The regex allows attributes on the element — openpyxl writes `<dc:title xmlns:dc="...">`,
    not `<dc:title>`, and a pattern without that allowance silently reports every fixture as
    having no title. (It did, on the first run of this check.)"""
    import re
    m = re.search(rf"<{tag}[^>]*>(.*?)</{tag}>", xml, re.S)
    return m.group(1) if m else None


@pytest.mark.parametrize("name,tag,present", [
    ("no-document-title", "dc:title", False),
    ("document-title-ok", "dc:title", True),
    ("no-document-language", "dc:language", False),
    ("document-language-ok", "dc:language", True),
])
def test_the_engine_fixtures_carry_or_withhold_the_right_property(corpus, name, tag, present):
    """What this file CAN prove without the .NET analyser, and the half that actually rots.

    Detection for 2.4.2 and 3.1.1 is asserted below, gated on the engine. But a fixture silently
    losing the property it was built around is a corpus defect that no engine is needed to catch
    — and it is the likelier failure, because it happens whenever someone edits the base workbook
    rather than the fixture. Splitting the two means a broken fixture fails everywhere, and only
    the detection claim waits for CI."""
    out, rows = corpus
    xml = _core_xml(out / rows[name]["file"])
    value = _prop(xml, tag)
    if present:
        assert value, f"{name} should carry a non-empty <{tag}> and has {value!r}"
    else:
        assert not value, f"{name} deliberately withholds <{tag}> but carries {value!r}"


def test_every_other_fixture_declares_both_a_title_and_a_language(corpus):
    """The correctness fix this pair of criteria surfaced, pinned so it cannot regress.

    openpyxl leaves `properties.language` at None, and XLSX-LANG-001 reports 3.1.1 for exactly
    that — so before the base workbook set one, EVERY fixture here also carried a 3.1.1 finding
    under the engine. The single-criterion labels were true only because a bare container has no
    .NET, which is the worst way for a ground-truth corpus to be right."""
    out, rows = corpus
    withholding = {"no-document-title": "dc:title", "no-document-language": "dc:language"}
    for name, row in rows.items():
        xml = _core_xml(out / row["file"])
        for tag in ("dc:title", "dc:language"):
            if withholding.get(name) == tag:
                continue
            assert _prop(xml, tag), (
                f"{name} has no <{tag}> — under the .NET analyser it would also raise "
                f"{'2.4.2' if tag == 'dc:title' else '3.1.1'}, so its label is wrong in CI")


@pytest.mark.skipif(not OFFICE_OK, reason=NO_OFFICE)
@pytest.mark.parametrize("name,sc,fires", [
    ("no-document-title", "2.4.2", True),
    ("document-title-ok", "2.4.2", False),
    ("no-document-language", "3.1.1", True),
    ("document-language-ok", "3.1.1", False),
])
def test_the_engine_confirms_the_declared_pairs(corpus, name, sc, fires):
    """The detection half — and the reason these two pairs sit in DECLARED_ENGINE rather than
    DECLARED. No first-party Python detector exists for 2.4.2 or 3.1.1 on ANY Office format, so
    unlike every other pair in this corpus the label is proven where the analyser is built and
    skipped where it is not.

    That is a weaker guarantee, and it is still worth having: both are among the seventeen
    (criterion, format) pairs in the preset that can return a PASS, so before these fixtures a
    clean scan CERTIFIED the file against a criterion nothing in the suite checked.

    SC ids come through `assessment_policy._extract_sc`, not a string split. The .NET analyser
    reports `wcag` in enum form — "SC_2_4_2" — where the first-party checks report
    "2.4.2 Page Titled", so splitting on whitespace yields "SC_2_4_2" and matches nothing. That
    is what the first CI run of this test failed on: both detectors HAD fired, and the assertion
    could not see it. `_extract_sc` is the repo's own normaliser for this exact spread of formats
    and is what `proposals.verify_residual_scs` uses, so the ids line up with the scan traces."""
    from assessment_policy import _extract_sc
    from scanner import analyse_and_assess
    out, rows = corpus
    path = out / rows[name]["file"]
    fd, _ = analyse_and_assess(path.parent, path.name, detect_pii=False)
    # NOT a walrus in a comprehension. PEP 572 binds an assignment expression in the CONTAINING
    # scope, so `{sc for i in ... if (sc := ...)}` rebinds the parametrised `sc` to whichever
    # criterion the last issue happened to carry — and every assertion below then judged that
    # leaked value. In the `fires=True` direction it asserted a value just taken out of `found`
    # was in `found`, which is true by construction, so those rows passed VACUOUSLY.
    #
    # Found on the pptx copy of this test (#1390), where a `fires=False` row happened to leak a
    # criterion that WAS present and failed with the give-away message "table-header-ok is the
    # clean control for 1.1.1" — on a test parametrised with 1.3.1.
    found = {s for i in (fd or {}).get("issues", []) if (s := _extract_sc(i.get("wcag", "")))}
    if fires:
        assert sc in found, (
            f"{name} declares {sc} but the analyser reported {sorted(found) or 'nothing'}")
    else:
        assert sc not in found, f"{name} is the clean control for {sc} but the analyser flagged it"

    # The other engine pair must stay quiet whichever way this fixture goes — that is the
    # correctness fix proving itself against the real analyser. `no-document-title` withholds a
    # title but HAS a language, so it must raise 2.4.2 and not 3.1.1; before the base workbook
    # stamped a language it would have raised both, and every other fixture here would have
    # raised 3.1.1 as well. Asserted in CI because that is the only place it can be.
    other = "3.1.1" if sc == "2.4.2" else "2.4.2"
    assert other not in found, (
        f"{name} also raised {other} — the base workbook has stopped stamping "
        f"{'a language' if other == '3.1.1' else 'a title'}, so every fixture in this corpus is "
        f"now carrying an undeclared finding")


# ── 1.4.5 is read out of the pixels, so it needs its own lane and its own guard ──

def test_ocr_is_present_in_ci():
    """The environment-conditional skip on the 1.4.5 rows must be a FALLBACK, never the normal
    state. Both ci.yml and azure-pipelines.yml run scripts/install_tesseract.sh, and the .docx
    gate makes the same assertion for the same reason: a skip nobody notices is one edit away
    from being how a criterion stops being covered.

    Skipped OFF CI so a bare developer checkout is not failed for a dependency it never claimed
    to have."""
    import os
    if not (os.environ.get("CI") or os.environ.get("TF_BUILD")):
        pytest.skip("not CI — tesseract is optional on a developer checkout")
    assert _ocr.is_available(), (
        "tesseract is unavailable in CI, so 1.4.5 was NOT exercised — the corpus would report "
        "the pair as covered while proving nothing about it")


# ── 1.3.1 and 1.3.2: the structural half, and the sweep that found two mislabels ─────────────
# 1.3.2 is raised by THREE rules, so the sweep checks all three. A fixture only has to trip one of
# them to carry an undeclared finding, and two already did before this pair was declared.

MERGED_CELL_THRESHOLD = 20   # Xlsx/Rules/MergedCellsRule.cs, `MergedCellThreshold`


def _table_without_header(path: Path) -> str | None:
    """TableHeaderRule: a table PART whose headerRowCount is an explicit 0 (absent means 1)."""
    import re
    import zipfile
    with zipfile.ZipFile(str(path)) as z:
        for name in z.namelist():
            if not re.match(r"xl/tables/table\d+\.xml$", name):
                continue
            match = re.search(r'headerRowCount="(\d+)"', z.read(name).decode("utf-8", "replace"))
            if match and int(match.group(1)) == 0:
                return f"{name}: headerRowCount=0"
    return None


def _hidden_row_with_data(path: Path) -> str | None:
    """HiddenContentRule: a hidden row or column that still holds non-blank data."""
    from openpyxl import load_workbook
    for ws in load_workbook(str(path)).worksheets:
        for dim in ws.row_dimensions.values():
            if dim.hidden and any(c.value not in (None, "") for c in ws[dim.index]):
                return f"{ws.title}: row {dim.index} hidden with data"
        for key, dim in ws.column_dimensions.items():
            if not dim.hidden:
                continue
            for row in ws.iter_rows(min_col=dim.min or 1, max_col=dim.max or 1):
                if any(c.value not in (None, "") for c in row):
                    return f"{ws.title}: column {key} hidden with data"
    return None


def _too_many_merges(path: Path) -> str | None:
    """MergedCellsRule: STRICTLY more than the threshold, which is why this reads `>`."""
    import re
    import zipfile
    with zipfile.ZipFile(str(path)) as z:
        for name in z.namelist():
            if not re.match(r"xl/worksheets/sheet\d+\.xml$", name):
                continue
            match = re.search(r'<mergeCells count="(\d+)"',
                              z.read(name).decode("utf-8", "replace"))
            if match and int(match.group(1)) > MERGED_CELL_THRESHOLD:
                return f"{name}: {match.group(1)} merged ranges"
    return None


def _blank_visible_sheet(path: Path) -> str | None:
    """BlankWorksheetRule: a VISIBLE sheet with no cell content and no anchored drawing.

    Hidden sheets are exempt by the rule's own comment — there is no navigation experience to
    flag them for.
    """
    from openpyxl import load_workbook
    for ws in load_workbook(str(path)).worksheets:
        if ws.sheet_state != "visible":
            continue
        if any(c.value not in (None, "") for row in ws.iter_rows() for c in row):
            continue
        if getattr(ws, "_images", None) or getattr(ws, "_charts", None):
            continue
        return f"{ws.title}: visible, no cells and no drawing"
    return None


def _first(*predicates):
    def check(path):
        for predicate in predicates:
            reason = predicate(path)
            if reason:
                return f"{predicate.__name__}: {reason}"
        return None
    return check


# Every engine predicate, keyed by the criterion it raises. Walked by the sweep below rather than
# hand-listed, so a fifth engine pair extends the sweep by construction.
ENGINE_PREDICATES = {
    "1.3.1": _table_without_header,
    "1.3.2": _first(_hidden_row_with_data, _too_many_merges, _blank_visible_sheet),
    "2.4.2": lambda p: None if _prop(_core_xml(p), "dc:title") else "no dc:title",
    "3.1.1": lambda p: None if _prop(_core_xml(p), "dc:language") else "no dc:language",
}


def test_the_sweep_covers_every_engine_declared_pair():
    """ANTI-VACUOUS ON THE SWEEP ITSELF. A predicate map that fell behind DECLARED_ENGINE would
    let the next engine pair be added with no undeclared-finding check, and the sweep below would
    still pass — silently checking three criteria out of four."""
    assert set(ENGINE_PREDICATES) == set(gen.DECLARED_ENGINE)


def test_no_fixture_carries_an_undeclared_engine_finding(corpus):
    """THE SWEEP THAT FOUND TWO MISLABELS, and one of them was a CONTROL.

    `sheet-tabs-default` and `sheet-tabs-named-ok` each created bare extra sheets, and
    BlankWorksheetRule raises 1.3.2 on any visible empty sheet — so under the analyser both
    carried an undeclared 1.3.2 alongside the 2.4.6 they declare. The adversarial one is the
    worse half: a control that fires is not a control, and it would have been a false positive
    on the day 1.3.2 was declared.

    Nothing caught it before, because 1.3.2 was undeclared and no first-party detector reports
    it. Running the sweep BEFORE declaring the pair is what made it a fixture fix rather than a
    CI failure blamed on the detector.
    """
    out, rows = corpus
    for name, row in rows.items():
        path = out / row["file"]
        for sc, predicate in ENGINE_PREDICATES.items():
            reason = predicate(path)
            if reason:
                assert sc in row["expect"], (
                    f"{name} would raise {sc} under the analyser ({reason}) and does not declare "
                    f"it — its label is wrong in CI")


@pytest.mark.parametrize("name,fires", [
    ("table-no-header", True),
    ("table-header-ok", False),
])
def test_the_table_fixtures_carry_or_withhold_what_the_rule_reads(corpus, name, fires):
    out, rows = corpus
    reason = _table_without_header(out / rows[name]["file"])
    if fires:
        assert reason, f"{name} is the 1.3.1 fixture and its table part now declares a header row"
    else:
        assert reason is None, f"{name} should declare a header row and does not"


def test_the_table_fixtures_write_a_real_table_part(corpus):
    """A BLOCK OF CELLS THAT LOOKS LIKE A TABLE IS NOT ONE. TableHeaderRule walks
    `TableDefinitionParts`; a fixture that formatted a range instead would declare the pair and
    detect nothing, which is the failure this corpus exists to prevent."""
    import re
    import zipfile
    out, rows = corpus
    for name in ("table-no-header", "table-header-ok"):
        with zipfile.ZipFile(str(out / rows[name]["file"])) as z:
            assert [n for n in z.namelist() if re.match(r"xl/tables/table\d+\.xml$", n)], (
                f"{name} has no table part — the rule has nothing to read")


@pytest.mark.parametrize("name,fires", [
    ("hidden-row", True),
    ("hidden-row-ok", False),
])
def test_the_hidden_row_fixtures_carry_or_withhold_what_the_rule_reads(corpus, name, fires):
    out, rows = corpus
    reason = _hidden_row_with_data(out / rows[name]["file"])
    if fires:
        assert reason, f"{name} is the 1.3.2 fixture and its row is no longer hidden-with-data"
    else:
        assert reason is None, f"{name} should have no hidden row and does: {reason}"


def test_the_hidden_row_pair_differs_only_in_the_hidden_flag(corpus):
    """Same cells, same values; one row hidden. If 1.3.2 ever reported on the PRESENCE of rows
    rather than on their visibility, the control would fire too."""
    from openpyxl import load_workbook
    out, rows = corpus

    def cells(name):
        ws = load_workbook(str(out / rows[name]["file"])).worksheets[0]
        return [(c.coordinate, c.value) for row in ws.iter_rows() for c in row
                if c.value not in (None, "")]

    assert cells("hidden-row") == cells("hidden-row-ok")


def test_a_hidden_row_with_no_data_is_not_the_trigger(corpus):
    """The rule needs hidden AND non-blank. Built here rather than asserted, because "hidden rows
    are flagged" is the plausible misreading that would make the fixture's label wrong."""
    from openpyxl import load_workbook
    out, _rows = corpus
    path = out / "docs" / "_hidden-but-empty.xlsx"
    wb = load_workbook(str(out / "docs" / "hidden-row-ok.xlsx"))
    ws = wb.worksheets[0]
    ws.row_dimensions[9].hidden = True          # hidden, and holds nothing
    wb.save(path)
    assert _hidden_row_with_data(path) is None
    path.unlink()


@pytest.mark.skipif(not OFFICE_OK, reason=NO_OFFICE)
@pytest.mark.parametrize("name,sc,fires", [
    ("table-no-header", "1.3.1", True),
    ("table-header-ok", "1.3.1", False),
    ("hidden-row", "1.3.2", True),
    ("hidden-row-ok", "1.3.2", False),
])
def test_the_engine_confirms_the_structure_pairs(corpus, name, sc, fires):
    """The detection half, gated on the analyser being built. Same `_extract_sc` normalisation as
    the title/language pair: the .NET analyser reports `wcag` in enum form ("SC_1_3_1") where the
    first-party checks report "1.3.1 Info and Relationships"."""
    from assessment_policy import _extract_sc
    from scanner import analyse_and_assess
    out, rows = corpus
    path = out / rows[name]["file"]
    fd, _ = analyse_and_assess(path.parent, path.name, detect_pii=False)
    found = {s for i in (fd or {}).get("issues", []) if (s := _extract_sc(i.get("wcag", "")))}
    if fires:
        assert sc in found, (
            f"{name} declares {sc} but the analyser reported {sorted(found) or 'nothing'}")
    else:
        assert sc not in found, f"{name} is the clean control for {sc} but the analyser flagged it"

    for other in set(gen.DECLARED_ENGINE) - {sc}:
        if other in rows[name]["expect"]:
            continue
        assert other not in found, (
            f"{name} also raised {other}, which it does not declare — the base workbook has "
            f"stopped supplying something every fixture relied on")
