"""Renaming a sheet or column must rename every REFERENCE to it (WCAG 2.4.6 write-back).

THE DEFECT THIS FILE PINS, found while trying to prove the 2.4.6 xlsx lane end to end. A sheet's
name is not stored once. `xl/workbook.xml` holds the tab; every formula, defined name and chart
series that points at the sheet spells the name out again in its own part. `apply_xlsx_labels`
renamed the tab and nothing else, so a workbook that was fine before remediation opened with
`#REF!` afterwards:

    xl/workbook.xml            <sheet name="Findings by quarter" …/>     ← renamed
    xl/worksheets/sheet2.xml   <f>Sheet1!B1</f>                          ← dangling
    xl/worksheets/sheet2.xml   <f>SUM(Sheet1!B1:B1)</f>                  ← dangling
    xl/workbook.xml            <definedName>Sheet1!$B$1</definedName>    ← dangling

WHY IT SURVIVED. It is invisible from every direction the pipeline looks. The criterion clears —
no default-named tabs remain — so a real re-scan reports 2.4.6 gone and the lane credits the
approved value. No detector reads formulas, so nothing downstream notices either. The
remediation WAS the damage, unattended, exactly as the 1.4.3 PDF contrast fixer was: it rewrote
compliant dark-theme PDFs from 21:1 to 3.66:1 and reported success.

The tests below are grouped by what would break without them, and each is written to fail
against the old writer: the reference cases were all dangling, and the two "must not touch"
cases are the ways a careless fix corrupts a workbook it was asked to repair.
"""
from __future__ import annotations

import io
import re
import sys
import tempfile
import zipfile
from html import unescape
from pathlib import Path

import pytest

ACP = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ACP / "api"))

pytest.importorskip("openpyxl")

OLD = "Sheet1"
NEW = "Findings by quarter"      # contains spaces, so every reference to it must be QUOTED
FILE = "workbook.xlsx"


def _book(formulas: dict[str, str] | None = None, *, defined: str | None = None) -> bytes:
    """Three sheets; `formulas` are written into Sheet2, keyed by cell."""
    import openpyxl
    from openpyxl.workbook.defined_name import DefinedName

    wb = openpyxl.Workbook()
    wb.active.title = OLD
    wb.create_sheet("Sheet2")
    wb.create_sheet("Sheet3")
    wb[OLD]["A1"] = "Quarter"
    wb[OLD]["B1"] = 128
    for ref, f in (formulas or {}).items():
        wb["Sheet2"][ref] = f
    if defined:
        wb.defined_names.add(DefinedName("Findings", attr_text=defined))
    out = Path(tempfile.mkdtemp()) / FILE
    wb.save(out)
    return out.read_bytes()


def _rename(data: bytes, values: dict[str, str]) -> bytes:
    from apply_xlsx_labels import apply_xlsx_labels
    fixed, applied, unresolved = apply_xlsx_labels(data, values)
    assert applied and not unresolved, (applied, unresolved)
    return fixed


def _formulas(data: bytes) -> list[str]:
    """Every stored formula in the package, read out of the XML rather than through a library
    that might normalise what it found.

    XML entities are decoded so the assertions read as formulas: a comparison is stored as
    `IF(A1&gt;0,…)`, and asserting against the escaped form would be asserting about XML
    encoding rather than about the reference the rename was supposed to follow.
    """
    out: list[str] = []
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        for name in sorted(z.namelist()):
            if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name):
                out += re.findall(r"<f\b[^>]*>([^<]*)</f>", z.read(name).decode("utf-8"))
    return [unescape(f) for f in out]


def _defined_names(data: bytes) -> list[str]:
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        xml = z.read("xl/workbook.xml").decode("utf-8")
    return re.findall(r"<definedName\b[^>]*>([^<]*)</definedName>", xml)


def _tabs(data: bytes) -> list[str]:
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        xml = z.read("xl/workbook.xml").decode("utf-8")
    return re.findall(r'<sheet\b[^>]*\bname="([^"]*)"', xml)


# ── 1. the tab really is renamed — the precondition for everything below ──────

def test_the_tab_is_renamed():
    assert _tabs(_rename(_book(), {f"sheet:{OLD}": NEW})) == [NEW, "Sheet2", "Sheet3"]


# ── 2. no reference is left naming a sheet that no longer exists ──────────────

@pytest.mark.parametrize("before,after", [
    ("=Sheet1!B1", f"'{NEW}'!B1"),
    ("=SUM(Sheet1!B1:B1)", f"SUM('{NEW}'!B1:B1)"),
    ("=IF(Sheet1!B1>0,Sheet1!A1,0)", f"IF('{NEW}'!B1>0,'{NEW}'!A1,0)"),
    ("=SUM(Sheet1:Sheet3!B1)", f"SUM('{NEW}:Sheet3'!B1)"),        # 3-D range
])
def test_every_formula_reference_follows_the_rename(before, after):
    """Each of these was left dangling by the old writer. The replacement is emitted QUOTED
    because the new name has spaces — `Findings by quarter!B1` is not a valid reference."""
    fixed = _rename(_book({"B1": before}), {f"sheet:{OLD}": NEW})
    assert _formulas(fixed) == [after]
    assert OLD not in _formulas(fixed)[0], "a reference still names the old sheet"


def test_a_defined_name_follows_the_rename():
    fixed = _rename(_book(defined=f"{OLD}!$B$1"), {f"sheet:{OLD}": NEW})
    assert _defined_names(fixed) == [f"'{NEW}'!$B$1"]


def test_the_workbook_reopens_with_every_sheet_intact():
    """Through openpyxl, which had no part in the write."""
    import openpyxl
    fixed = _rename(_book({"B1": "=Sheet1!B1"}), {f"sheet:{OLD}": NEW})
    p = Path(tempfile.mkdtemp()) / FILE
    p.write_bytes(fixed)
    wb = openpyxl.load_workbook(str(p))
    assert wb.sheetnames == [NEW, "Sheet2", "Sheet3"]
    assert wb[NEW]["A1"].value == "Quarter" and wb[NEW]["B1"].value == 128


# ── 3. what the rewrite must NOT touch ────────────────────────────────────────
#
# A rename that rewrites too much corrupts the workbook just as surely as one that rewrites too
# little, and it is the easier mistake to make with a regex.

def test_a_different_sheet_whose_name_merely_contains_the_old_one_is_untouched():
    """`MySheet1!B1` and `Sheet10!B1` both contain "Sheet1" as a substring and reference other
    sheets entirely. A naive replace renames them and breaks two more formulas."""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = OLD
    wb.create_sheet("MySheet1")
    wb.create_sheet("Sheet10")
    wb.create_sheet("Sheet2")
    wb["Sheet2"]["B1"] = "=MySheet1!B1"
    wb["Sheet2"]["B2"] = "=Sheet10!B1"
    wb["Sheet2"]["B3"] = "=Sheet1!B1"
    o = Path(tempfile.mkdtemp()) / FILE
    wb.save(o)

    fixed = _rename(o.read_bytes(), {f"sheet:{OLD}": NEW})
    assert _formulas(fixed) == ["MySheet1!B1", "Sheet10!B1", f"'{NEW}'!B1"]


def test_text_that_merely_looks_like_a_reference_is_untouched():
    """`="Sheet1!B1"` is a string a user typed. It is not a reference and rewriting it silently
    changes their data — the failure a rewriter gets wrong in the opposite direction."""
    fixed = _rename(_book({"B1": '="Sheet1!B1 is the source"', "B2": "=Sheet1!B1"}),
                    {f"sheet:{OLD}": NEW})
    assert _formulas(fixed) == ['"Sheet1!B1 is the source"', f"'{NEW}'!B1"]


def test_a_rename_that_does_not_apply_rewrites_nothing():
    """An approved value naming a sheet this workbook does not have. Nothing is renamed, so
    nothing may be rewritten either — the references still point at real sheets."""
    from apply_xlsx_labels import apply_xlsx_labels
    data = _book({"B1": "=Sheet1!B1"})
    fixed, applied, unresolved = apply_xlsx_labels(data, {"sheet:Nonexistent": NEW})
    assert applied == [] and unresolved == ["sheet:Nonexistent"]
    assert _formulas(fixed) == ["Sheet1!B1"]
    assert _tabs(fixed) == [OLD, "Sheet2", "Sheet3"]


# ── 4. the same defect on the table-column half of the lane ───────────────────

def _table_book() -> bytes:
    """A worksheet table with a default column name, plus formulas referencing it."""
    import openpyxl
    from openpyxl.worksheet.table import Table

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"], ws["B1"] = "Column1", "Column2"
    ws["A2"], ws["B2"] = 1, 2
    ws.add_table(Table(displayName="Sales", ref="A1:B2"))
    ws["D1"] = "=SUM(Sales[Column1])"
    ws["D2"] = '="Sales[Column1] is the header"'
    o = Path(tempfile.mkdtemp()) / FILE
    wb.save(o)
    return o.read_bytes()


def test_a_structured_column_reference_follows_a_column_rename():
    """`Sales[Column1]` names the column the same way `Sheet1!A1` names the sheet, and broke
    the same way."""
    fixed = _rename(_table_book(), {"table:Sales#col:Column1": "Region"})
    assert "SUM(Sales[Region])" in _formulas(fixed)


def test_text_that_looks_like_a_structured_reference_is_untouched():
    fixed = _rename(_table_book(), {"table:Sales#col:Column1": "Region"})
    assert '"Sales[Column1] is the header"' in _formulas(fixed)
