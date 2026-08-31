"""The 2.4.6 xlsx structure-label lane, proved end to end (WCAG 2.4.6 Headings and Labels).

Same bar as the lanes before: the original workbook trips the finding, an approval changes the
saved workbook, a REAL re-scan verifies it, unrelated content survives, and a broken engine
earns no credit. Nothing but the blob store is patched.

WHY THIS ONE LANDED LAST, and it is the reason worth reading. The proof was written first and
then withheld: the lane cleared its criterion by renaming the sheet tab in `xl/workbook.xml` and
nothing else, leaving every formula, defined name and chart series that referenced the sheet
naming one that no longer existed — `#REF!` in Excel. It was invisible from every direction the
pipeline looks (no default tabs remain, so a real re-scan reports 2.4.6 gone and the lane
credits the value; no detector reads formulas). Registering the lane on that evidence would have
certified the damage. #1076 fixed the writer; `tests/test_xlsx_label_references.py` pins it, and
`test_the_rename_does_not_strand_a_formula_that_referenced_the_sheet` below re-asserts the same
guarantee THROUGH THE LANE rather than through the writer alone — because that is the level a
customer downloading the file experiences.

WHERE THE DETECTOR'S GATE IS, which shapes both the fixture and the partial-write control:
`office_structure` fires when a workbook has **two or more** default-named sheet tabs
(`Sheet1`, `Sheet2`, …) or any default-named table column. Two consequences that a
one-fixture-fits-all test would get wrong:

  * A two-sheet workbook with ONE tab renamed already has fewer than two defaults left, so the
    criterion clears — a "partial write is not credited" control built that way proves nothing.
    The control here uses THREE default sheets and renames one, which leaves two and still
    fails. That number is not arbitrary; it was measured.
  * A single default tab is not a finding at all. `test_one_default_tab_is_not_a_finding` pins
    that, so the fixture's second sheet is understood as load-bearing rather than decorative.

WHAT THIS CLAIMS: the tabs a reader sees now carry meaningful names, every reference to them
still resolves, the workbook survived, and ACP's own criterion stops firing. NOT that the names
are the ones a human would choose — naming is a judgement, which is why the lane is `assisted`.
"""
from __future__ import annotations

import io
import re
import sys
import tempfile
import zipfile
from html import unescape
from pathlib import Path

import pytest

ACP = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ACP / "api"))

pytest.importorskip("openpyxl")

FILE = "quarterly.xlsx"
SID = "rv-xlsx-246"

APPROVED = "Findings by quarter"       # prose, with spaces — so references to it must be QUOTED
SECOND = "Owners"
SUMMARY = "Summary"                    # a sheet that is already meaningfully named


def _book(*, defaults: int = 2, formula: bool = True) -> bytes:
    """`defaults` default-named sheets plus one meaningfully named sheet holding a cross-sheet
    formula, so every fixture here can tell a rename that preserved the workbook from one that
    silently broke it."""
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "Sheet1"
    for n in range(2, defaults + 1):
        wb.create_sheet(f"Sheet{n}")
    wb.create_sheet(SUMMARY)

    wb["Sheet1"]["A1"] = "Quarter"
    wb["Sheet1"]["B1"] = 128
    wb[SUMMARY]["A1"] = "Total findings"
    if formula:
        wb[SUMMARY]["B1"] = "=Sheet1!B1"

    out = Path(tempfile.mkdtemp()) / FILE
    wb.save(out)
    return out.read_bytes()


def _assess(data: bytes) -> set[str]:
    from assessment_policy import _extract_sc
    from scanner import analyse_and_assess
    with tempfile.TemporaryDirectory() as d:
        (Path(d) / FILE).write_bytes(data)
        fd, _ = analyse_and_assess(Path(d), FILE, detect_pii=False)
    return {sc for i in (fd or {}).get("issues", []) if (sc := _extract_sc(i.get("wcag", "")))}


def _spill(data: bytes) -> Path:
    p = Path(tempfile.mkdtemp()) / FILE
    p.write_bytes(data)
    return p


def _tabs(data: bytes) -> list[str]:
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        xml = z.read("xl/workbook.xml").decode("utf-8")
    return re.findall(r'<sheet\b[^>]*\bname="([^"]*)"', xml)


def _formulas(data: bytes) -> list[str]:
    out: list[str] = []
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        for name in sorted(z.namelist()):
            if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name):
                out += re.findall(r"<f\b[^>]*>([^<]*)</f>", z.read(name).decode("utf-8"))
    return [unescape(f) for f in out]


class _Blob:
    def __init__(self, data: bytes):
        self.data, self.uploads = data, []

    def download_remediated(self, owner, sid, f):
        return self.data

    def upload_remediated(self, owner, sid, f, data, mime):
        self.data = data
        self.uploads.append((f, mime))
        return "http://b/2"


@pytest.fixture()
def store(monkeypatch):
    import store as store_mod
    monkeypatch.setattr(store_mod, "_SQLITE_PATH", Path(tempfile.mkdtemp()) / "rv.db")
    return store_mod.Store()


def _seed(store, values: dict[str, str]) -> str:
    store.init_scan_run(SID, "drive", 1, "2026-08-31T00:00:00Z", "rubric", "hash")
    store.save_file_result(SID, {
        "file": FILE, "engine": "office", "status": "pass", "score": 60, "compliant": 0,
        "skipped_rules": 0, "drive_file_id": "d1",
        "issues": [{"ruleId": "XLSX_DEFAULT_LABELS", "wcag": "2.4.6 Headings and Labels",
                    "severity": "MODERATE"}],
    }, "2026-08-31T00:00:00Z")
    store.record_remediation(SID, FILE, drive_write_url="http://d/1", blob_url="http://b/1")
    item_id = store.enqueue_proposals(SID, FILE, "2.4.6", [
        {"locator": loc, "before": loc.split(":", 1)[1], "proposed_value": "", "rationale": "r",
         "source": "reviewer"} for loc in values], rule_name="Headings and Labels")
    store.update_hitl_item(item_id, "approved", None, None)
    store.approve_proposal_values(item_id, list(values.values()))
    return item_id


def _run_lane(monkeypatch, store, blob):
    import core
    import handlers
    monkeypatch.setattr(core, "store", store)
    monkeypatch.setitem(sys.modules, "blob", blob)
    handlers._apply_approved_values({"scan_id": SID, "file": FILE}, {})


# ── 1. the finding and its gate ───────────────────────────────────────────────

def test_a_real_assessment_reports_2_4_6_on_default_sheet_names():
    assert "2.4.6" in _assess(_book())


def test_the_first_party_gate_is_two_default_tabs():
    """Why the partial-write control below needs THREE default sheets, asked of the FIRST-PARTY
    detector rather than of a full scan — and that distinction is the whole point.

    `office_structure` fires at TWO or more default tabs, so renaming one of two drops below its
    gate and clears the criterion: a control built that way would prove nothing. Three leaves
    two, and still fails.

    THE .NET ANALYSER IS STRICTER, and asking a full scan would have conflated the two. Its
    `SheetNameRule` flags EVERY sheet matching `^Sheet\d+$` individually, with no count gate at
    all — so on a host where the Office CLI is built, even one default tab is a 2.4.6 finding.
    The first draft of this test asserted `"2.4.6" not in _assess(_book(defaults=1))`, passed
    locally where no analyser is installed, and failed CI with
    `assert '2.4.6' not in {'2.4.2', '2.4.6', '3.1.1'}` — the same environment-dependence that
    turned #1069's shard 3 red, in a new place.

    Three sheets is therefore the right fixture under BOTH engines: it is above the first-party
    gate, and .NET is only ever stricter. Asking the first-party detector directly makes this
    assertion say what it means and hold everywhere.
    """
    from office_structure import xlsx_structure_checks
    def first_party_scs(data: bytes) -> set[str]:
        return {f["wcag"].split()[0] for f in xlsx_structure_checks(_spill(data))
                if f.get("wcag")}

    assert "2.4.6" not in first_party_scs(_book(defaults=1))
    assert "2.4.6" in first_party_scs(_book(defaults=2))
    assert "2.4.6" in first_party_scs(_book(defaults=3))


def test_a_meaningfully_named_workbook_is_not_flagged():
    """The control. Without it, a detector that flagged every workbook would satisfy the first
    test and nothing would notice."""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active.title = APPROVED
    wb.create_sheet(SECOND)
    out = Path(tempfile.mkdtemp()) / FILE
    wb.save(out)
    assert "2.4.6" not in _assess(out.read_bytes())


# ── 2. approval → write → re-scan → credit, through the real path ─────────────

@pytest.fixture()
def applied(store, monkeypatch):
    original = _book()
    blob = _Blob(original)
    _seed(store, {"sheet:Sheet1": APPROVED, "sheet:Sheet2": SECOND})
    _run_lane(monkeypatch, store, blob)
    return blob, store, original


def test_the_saved_workbook_carries_the_approved_names(applied):
    blob, _, _ = applied
    assert _tabs(blob.data) == [APPROVED, SECOND, SUMMARY]


def test_the_rename_does_not_strand_a_formula_that_referenced_the_sheet(applied):
    """The assertion this lane was withheld for, made THROUGH THE LANE rather than through the
    writer alone. A sheet's name is spelled out again in every formula that points at it, and
    renaming only the tab left `=Sheet1!B1` naming a sheet that no longer existed — `#REF!` in
    the workbook a customer downloads, while 2.4.6 read as cleared.

    The replacement is quoted because the approved name has spaces; an unquoted
    `Findings by quarter!B1` is not a valid reference either.
    """
    blob, _, _ = applied
    assert _formulas(blob.data) == [f"'{APPROVED}'!B1"]
    assert "Sheet1" not in _formulas(blob.data)[0], "a formula still names the old sheet"


def test_the_workbook_reopens_with_its_data_intact(applied):
    """Through openpyxl, which had no part in the write."""
    import openpyxl
    blob, _, _ = applied
    path = _spill(blob.data)
    assert zipfile.ZipFile(path).testzip() is None
    wb = openpyxl.load_workbook(str(path))
    assert wb.sheetnames == [APPROVED, SECOND, SUMMARY]
    assert wb[APPROVED]["A1"].value == "Quarter" and wb[APPROVED]["B1"].value == 128
    assert wb[SUMMARY]["A1"].value == "Total findings"


def test_the_already_named_sheet_is_untouched(applied):
    """Scoped write: only the default-named tabs were approved, and only they change."""
    blob, _, _ = applied
    assert SUMMARY in _tabs(blob.data)


def test_a_second_real_assessment_no_longer_reports_2_4_6(applied):
    """THE claim: a fresh assessment of the SAVED bytes, not the writer's return value."""
    blob, _, _ = applied
    assert "2.4.6" not in _assess(blob.data)


def test_the_row_is_credited_and_the_copy_is_stored(applied):
    blob, store, _ = applied
    assert store.count_unapplied_approved_values(SID, FILE) == 0
    assert blob.uploads


# ── 3. where the lane must NOT credit ─────────────────────────────────────────

def test_a_partial_write_is_not_credited_because_the_criterion_still_fails(store, monkeypatch):
    """THREE default sheets, ONE renamed. Two defaults remain, so 2.4.6 still fails and nothing
    may be credited.

    The count is measured, not chosen for symmetry with the other lanes' controls. On a
    two-sheet workbook renaming one tab leaves a single default, which is BELOW the detector's
    gate — the criterion would clear and this control would silently prove nothing.
    """
    from apply_xlsx_labels import apply_xlsx_labels
    original = _book(defaults=3)

    written, ap, _ = apply_xlsx_labels(original, {"sheet:Sheet1": APPROVED})
    assert ap, "the writer refused the value, so this control is not about crediting"
    assert _tabs(written)[0] == APPROVED
    assert "2.4.6" in _assess(written), (
        "renaming one of three default tabs cleared the criterion — this control cannot "
        "distinguish a withheld credit from a cleared one")

    blob = _Blob(original)
    _seed(store, {"sheet:Sheet1": APPROVED})
    _run_lane(monkeypatch, store, blob)

    assert store.count_unapplied_approved_values(SID, FILE) == 1
    assert store.mark_file_compliant_if_reviewed(SID, FILE) is False
    assert not blob.uploads
    assert blob.data == original


def test_renaming_a_tab_to_another_default_name_is_not_credited(store, monkeypatch):
    """The reviewer types "Sheet3". The write succeeds and the workbook is no better: a default
    name replaced by another default name, so the criterion still fails."""
    from apply_xlsx_labels import apply_xlsx_labels
    original = _book()

    written, ap, _ = apply_xlsx_labels(original, {"sheet:Sheet1": "Sheet3"})
    assert ap and "Sheet3" in _tabs(written)
    assert "2.4.6" in _assess(written), "'Sheet3' is not read as a default name; control vacuous"

    blob = _Blob(original)
    _seed(store, {"sheet:Sheet1": "Sheet3"})
    _run_lane(monkeypatch, store, blob)

    assert store.count_unapplied_approved_values(SID, FILE) == 1
    assert not blob.uploads


def test_an_approval_aimed_at_a_sheet_that_is_not_there_is_not_credited(store, monkeypatch):
    """Nothing resolves, so nothing is renamed — and, since #1076, nothing is rewritten either:
    a rename that did not happen must not drag references with it."""
    original = _book()
    blob = _Blob(original)
    _seed(store, {"sheet:Nonexistent": APPROVED})
    _run_lane(monkeypatch, store, blob)

    assert blob.data == original
    assert _formulas(original) == ["Sheet1!B1"]
    assert store.count_unapplied_approved_values(SID, FILE) == 1
    assert not blob.uploads


# ── 4. a broken engine earns nothing ──────────────────────────────────────────

@pytest.mark.parametrize("name,script,timeout", [
    ("cannot be launched", None, None),
    ("exits non-zero", "#!/bin/sh\necho boom >&2\nexit 9\n", None),
    ("hangs past the timeout", "#!/bin/sh\nsleep 30\n", "2"),
])
def test_a_broken_office_analyser_never_credits_this_lane_either(monkeypatch, name, script,
                                                                 timeout):
    """Re-asserted per lane rather than assumed to inherit: the fail-open #1058 closed lived in
    ONE shared seam, so a regression there takes every lane at once. 2.4.6 comes from
    office_structure, pure Python running after the .NET call, so the residual is a real set
    even with no analyser at all."""
    import stat as _stat

    import scanner
    if script is None:
        monkeypatch.setattr(scanner, "DOTNET", "/nonexistent/dotnet", raising=False)
    else:
        fake = Path(tempfile.mkdtemp()) / "dotnet"
        fake.write_text(script)
        fake.chmod(fake.stat().st_mode | _stat.S_IEXEC)
        monkeypatch.setattr(scanner, "DOTNET", str(fake), raising=False)
    if timeout:
        monkeypatch.setenv("ACP_OFFICE_CLI_TIMEOUT", timeout)

    from proposals import verify_residual_scs
    residual = verify_residual_scs(_book(), FILE)
    assert residual is not None, (
        f"an office CLI that {name} made the re-scan return None — every approved value on this "
        f"lane would be credited on a scan that never happened")
    assert "2.4.6" in residual
