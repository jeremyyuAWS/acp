"""Regression guard for XLSX LinkPurposeRule (XLSX-LINK-001, WCAG 2.4.4).

Detection lives in the vendored .NET analyser
(engine/office-analysers/DigitalA11y.Analysers.DotNet/Xlsx/Rules/LinkPurposeRule.cs).

Two distinct hyperlink mechanisms must both be covered: standard cell hyperlinks (read
from the worksheet's own <hyperlinks> element) and formula-driven =HYPERLINK(...) cells,
which never appear in that element at all -- a detector checking only cell.hyperlink
(or only <hyperlinks>) would miss every formula-based link entirely. This builds a
workbook covering both mechanisms, generic vs. descriptive text, and a formula link with
no friendly name (which displays the raw URL), and asserts the rule fires only where it
should. Needs the built .NET CLI to run, so it self-skips when the analysers didn't run
(CLI not built).
"""
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "api"))
import scanner  # noqa: E402


def _link_findings(fdict) -> list[dict]:
    return [i for i in fdict.get("issues", []) if i.get("ruleId") == "XLSX-LINK-001"]


def _locations(findings) -> list[str]:
    return [str(f.get("location")) for f in findings]


@pytest.fixture(scope="module")
def scanned(tmp_path_factory):
    from openpyxl import Workbook

    d = tmp_path_factory.mktemp("xlsx-link")
    wb = Workbook()
    ws = wb.active
    ws.title = "Links"

    # A1: standard cell hyperlink, generic display text -- must be flagged
    ws["A1"] = "click here"
    ws["A1"].hyperlink = "https://example.com/generic"

    # A2: standard cell hyperlink, descriptive display text -- must NOT be flagged
    ws["A2"] = "2026 benefits enrollment form"
    ws["A2"].hyperlink = "https://example.com/benefits"

    # A3: formula-driven link with a generic friendly name -- must be flagged.
    # Formula cells need a cached value for openpyxl to round-trip one without a real
    # Excel calc engine; write the display text as the cell's value alongside the formula.
    ws["A3"] = "=HYPERLINK(\"https://example.com/report\", \"click here\")"

    # A4: formula-driven link with NO friendly name -- Excel displays the raw URL itself,
    # so the cached value is the URL; must be flagged as a raw-URL failure.
    ws["A4"] = "=HYPERLINK(\"https://example.com/report2\")"

    # A5: plain cell, no hyperlink, no HYPERLINK formula -- must stay silent
    ws["A5"] = "Just a regular label"

    name = "link-purpose.xlsx"
    path = d / name
    wb.save(str(path))

    # openpyxl writes formula cells with the formula as the *value* and no cached <v>
    # result -- patch the raw XML so CellValueHelper has a display value to read, the
    # same way a real Excel save (which always caches results) would produce.
    import re
    import zipfile

    with zipfile.ZipFile(path) as z:
        sheet_name = next(n for n in z.namelist() if re.match(r"xl/worksheets/sheet\d+\.xml$", n))
        xml = z.read(sheet_name).decode("utf-8")
        others = {n: z.read(n) for n in z.namelist() if n != sheet_name}

    def _inject_cached_value(xml: str, cell_ref: str, display: str) -> str:
        pattern = rf'(<c r="{cell_ref}"[^>]*t="str"[^>]*>)(<f>.*?</f>)(</c>)'
        return re.sub(pattern, rf'\1\2<v>{display}</v>\3', xml)

    xml = _inject_cached_value(xml, "A3", "click here")
    xml = _inject_cached_value(xml, "A4", "https://example.com/report2")

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for n, data in others.items():
            z.writestr(n, data)
        z.writestr(sheet_name, xml)

    fdict, _ = scanner.analyse_and_assess(d, name, detect_pii=False)
    findings = _link_findings(fdict)
    return {
        "locations": _locations(findings),
        "count": len(findings),
        "engine_ran": any(str(i.get("wcag", "")).startswith("SC_") for i in fdict.get("issues", [])),
    }


def test_generic_standard_hyperlink_is_flagged(scanned):
    if not scanned["engine_ran"]:
        pytest.skip("the .NET office analysers did not run (CLI not built)")
    assert any(":cell:A1" in loc for loc in scanned["locations"]), \
        "a standard hyperlink with generic display text ('click here') was not flagged"


def test_descriptive_standard_hyperlink_is_not_flagged(scanned):
    if not scanned["engine_ran"]:
        pytest.skip("the .NET office analysers did not run (CLI not built)")
    assert not any(":cell:A2" in loc for loc in scanned["locations"]), \
        "a standard hyperlink with descriptive display text was incorrectly flagged"


def test_formula_driven_link_with_generic_text_is_flagged(scanned):
    """The core gap this rule exists to close: =HYPERLINK() cells never appear in the
    worksheet's own <hyperlinks> element, so a detector reading only that element would
    silently miss this one entirely."""
    if not scanned["engine_ran"]:
        pytest.skip("the .NET office analysers did not run (CLI not built)")
    assert any(":cell:A3" in loc for loc in scanned["locations"]), \
        "a formula-driven =HYPERLINK() link with generic display text was not flagged"


def test_formula_driven_link_with_no_friendly_name_is_flagged_as_raw_url(scanned):
    if not scanned["engine_ran"]:
        pytest.skip("the .NET office analysers did not run (CLI not built)")
    assert any(":cell:A4" in loc for loc in scanned["locations"]), \
        "a formula-driven link with no friendly name (displays the raw URL) was not flagged"


def test_plain_non_hyperlink_cell_is_not_flagged(scanned):
    if not scanned["engine_ran"]:
        pytest.skip("the .NET office analysers did not run (CLI not built)")
    assert not any(":cell:A5" in loc for loc in scanned["locations"]), \
        "a plain cell with no hyperlink was incorrectly flagged"


def test_exactly_three_links_are_flagged(scanned):
    if not scanned["engine_ran"]:
        pytest.skip("the .NET office analysers did not run (CLI not built)")
    assert scanned["count"] == 3, \
        f"expected exactly A1, A3, A4 to fire (3 findings), got {scanned['count']}: {scanned['locations']}"
