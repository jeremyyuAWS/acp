"""The 1.1.1 xlsx alt-text lane, proved end to end (WCAG 1.1.1 Non-text Content).

Same bar as the docx and pptx alt proofs: the original workbook trips the finding, an approval
changes the saved workbook, a REAL re-scan verifies it, unrelated content survives, and a broken
engine earns no credit. Nothing but the blob store is patched.

WHAT IS DIFFERENT ON xlsx, and it is the reason a third alt-text file is worth its weight.

  * THE IMAGE IS NOT IN THE SHEET. SpreadsheetML keeps pictures in a separate drawing part
    (`xl/drawings/drawingN.xml`) that the worksheet references through a relationship, so the
    locator is `xl/drawings/drawing1.xml#Image 1` — a part the docx and pptx lanes never touch,
    reached by a different branch of `apply_alt._ALT_TAG_FOR_PART` and named `cNvPr` rather
    than `wp:docPr` or `p:cNvPr`. That branch's pattern accepts the tag with OR without an
    `xdr:` prefix, and the tolerance is load-bearing rather than defensive: Excel writes
    `<xdr:cNvPr>`, openpyxl binds the drawing namespace as the default and writes it bare. A
    locator the writer cannot resolve makes the lane unreachable however good detection and
    writing are on their own, so `test_the_finding_carries_a_locator_the_writer_can_resolve`
    pins the xlsx spelling.

  * openpyxl GIVES EVERY PICTURE A NAME AND A PLACEHOLDER DESCRIPTION. The image below arrives
    with `descr="Picture"` already set — not empty. That is exactly the case a detector keying
    on "is the attribute present" would miss, and it is the common one: openpyxl is what most
    non-Excel tooling writes workbooks with. `test_a_placeholder_description_is_still_a_finding`
    asserts the detector reads it as undescribed, because if it did not, every assertion below
    would be about a document that never failed.

  * THE GRID IS THE THING THAT MUST SURVIVE. Alt text lives in the drawing part; the cell
    values, the shared-string table and the sheet name live elsewhere, and a writer that
    rebuilt the package rather than editing one part would lose them silently. They are read
    back through openpyxl, which had no part in the write.

WHAT THIS CLAIMS: a `descr` a screen reader will announce is present, the workbook survived, and
ACP's own criterion stops firing. Whether the sentence DESCRIBES the picture is a human
judgement no detector makes — which is why 1.1.1 resolves to REVIEW and never to PASS.
"""
from __future__ import annotations

import io
import re
import sys
import tempfile
import zipfile
from pathlib import Path

import pytest

ACP = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ACP / "api"))

pytest.importorskip("openpyxl")

FILE = "findings.xlsx"
SID = "rv-xlsx-111"
DRAWING = "xl/drawings/drawing1.xml"

SHEET = "Findings"
FIRST = "A bar chart of Q3 findings, grouped by severity."
SECOND = "A photograph of the accessibility working group."
# The grid content. Strings go through the shared-string table, so these also prove the write
# did not disturb it.
CELLS = {"A1": "Quarter", "B1": "Findings", "A2": "Q3", "A3": "Q4"}

_PNG = (b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x06\x00\x00"
        b"\x00\x1f\x15\xc4\x89\x00\x00\x00\nIDATx\x9cc\x00\x01\x00\x00\x05\x00\x01\r\n-\xb4"
        b"\x00\x00\x00\x00IEND\xaeB`\x82")


def _book(*, pictures: int = 2, alts: tuple[str | None, ...] = (None, None)) -> bytes:
    """A named sheet with real cell content and `pictures` images anchored in the grid.

    TWO pictures is not padding: it is what makes
    `test_a_partial_write_is_not_credited_because_the_criterion_still_fails` able to fail.
    """
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage

    img = Path(tempfile.mkdtemp()) / "i.png"
    img.write_bytes(_PNG)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    for ref, val in CELLS.items():
        ws[ref] = val
    ws["B2"] = 128
    ws["B3"] = 96
    for n in range(pictures):
        ws.add_image(XLImage(str(img)), f"{chr(ord('D') + 3 * n)}2")

    out = Path(tempfile.mkdtemp()) / FILE
    wb.save(out)
    data = out.read_bytes()
    return _preset_descriptions(data, alts) if any(alts) else data


def _preset_descriptions(data: bytes, alts: tuple[str | None, ...]) -> bytes:
    """Set `descr` on the drawing's pictures by editing the part directly.

    openpyxl offers no way to do this: `SpreadsheetDrawing._picture_frame` hard-codes
    `descr="Picture"` on every image it writes. The described CONTROL therefore has to be built
    some other way — and it must not be built with `apply_alt`, because the control's whole job
    is to establish, independently of the writer, that what clears the criterion is a real
    description. So this is a plain string edit on the XML, sharing no code with the lane.
    """
    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data)) as zin, zipfile.ZipFile(out, "w") as zout:
        for name in zin.namelist():
            payload = zin.read(name)
            if name == DRAWING:
                xml = payload.decode("utf-8")
                for n, alt in enumerate(alts, start=1):
                    if not alt:
                        continue
                    xml = re.sub(
                        rf'(<(?:xdr:)?cNvPr\b[^>]*\bname="Image {n}"[^>]*?\bdescr=")[^"]*(")',
                        lambda m, a=alt: m.group(1) + a + m.group(2), xml)
                payload = xml.encode("utf-8")
            zout.writestr(name, payload)
    return out.getvalue()


def _assess(data: bytes) -> set[str]:
    from assessment_policy import _extract_sc
    from scanner import analyse_and_assess
    with tempfile.TemporaryDirectory() as d:
        (Path(d) / FILE).write_bytes(data)
        fd, _ = analyse_and_assess(Path(d), FILE, detect_pii=False)
    return {sc for i in (fd or {}).get("issues", []) if (sc := _extract_sc(i.get("wcag", "")))}


def _spill(data: bytes) -> Path:
    p = Path(tempfile.mkdtemp()) / FILE
    p.write_bytes(data)
    return p


def _locators(data: bytes) -> list[str]:
    """Off the FIRST-PARTY detector, not off a full scan: `scanner._collapse_duplicate_alt`
    drops first-party 1.1.1 findings wherever the .NET engine also reported 1.1.1, and the
    engine's finding carries `page`/`location` rather than the `locator` an applier resolves.
    A full scan therefore answers differently depending on whether the Office CLI is built —
    the failure that turned #1069's shard 3 red."""
    from formats.xlsx.detectors import non_text_content
    return [f["locator"] for f in non_text_content.detect(_spill(data)) if f.get("locator")]


def _drawing_xml(data: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        return z.read(DRAWING).decode("utf-8")


class _Blob:
    def __init__(self, data: bytes):
        self.data, self.uploads = data, []

    def download_remediated(self, owner, sid, f):
        return self.data

    def upload_remediated(self, owner, sid, f, data, mime):
        self.data = data
        self.uploads.append((f, mime))
        return "http://b/2"


@pytest.fixture()
def store(monkeypatch):
    import store as store_mod
    monkeypatch.setattr(store_mod, "_SQLITE_PATH", Path(tempfile.mkdtemp()) / "rv.db")
    return store_mod.Store()


def _seed(store, values: dict[str, str]) -> str:
    store.init_scan_run(SID, "drive", 1, "2026-08-31T00:00:00Z", "rubric", "hash")
    store.save_file_result(SID, {
        "file": FILE, "engine": "office", "status": "pass", "score": 60, "compliant": 0,
        "skipped_rules": 0, "drive_file_id": "d1",
        "issues": [{"ruleId": "XLSX_IMAGE_NO_ALT", "wcag": "1.1.1 Non-text Content",
                    "severity": "CRITICAL", "locator": loc} for loc in values],
    }, "2026-08-31T00:00:00Z")
    store.record_remediation(SID, FILE, drive_write_url="http://d/1", blob_url="http://b/1")
    item_id = store.enqueue_proposals(SID, FILE, "1.1.1", [
        {"locator": loc, "before": "(no alt text)", "proposed_value": "", "rationale": "r",
         "source": "reviewer"} for loc in values], rule_name="Non-text Content")
    store.update_hitl_item(item_id, "approved", None, None)
    store.approve_proposal_values(item_id, list(values.values()))
    return item_id


def _run_lane(monkeypatch, store, blob):
    import core
    import handlers
    monkeypatch.setattr(core, "store", store)
    monkeypatch.setitem(sys.modules, "blob", blob)
    handlers._apply_approved_values({"scan_id": SID, "file": FILE}, {})


@pytest.fixture(scope="module")
def undescribed() -> bytes:
    return _book(pictures=2, alts=(None, None))


# ── 1. the finding ────────────────────────────────────────────────────────────

def test_a_real_assessment_reports_1_1_1_on_undescribed_images(undescribed):
    assert "1.1.1" in _assess(undescribed)


def test_a_placeholder_description_is_still_a_finding(undescribed):
    """openpyxl writes `descr="Picture"` on every image it adds, so the attribute IS present
    and says nothing. A detector reading presence rather than usefulness would call this
    workbook clean, and every assertion below would be about a document that never failed."""
    assert 'descr="Picture"' in _drawing_xml(undescribed), (
        "openpyxl no longer writes its placeholder description — this control now proves "
        "nothing about the interesting case")
    assert "1.1.1" in _assess(undescribed)


def test_the_same_workbook_with_both_images_described_is_not_flagged():
    """The control. Without it a detector that flagged every image would satisfy the tests
    above and nothing would notice."""
    assert "1.1.1" not in _assess(_book(pictures=2, alts=(FIRST, SECOND)))


def test_the_finding_carries_a_locator_the_writer_can_resolve(undescribed):
    """The xlsx spelling: a DRAWING part, not the worksheet. The image is not in the sheet at
    all — SpreadsheetML references it through a relationship — so this is a different branch of
    the applier's part table than either Office sibling uses."""
    from apply_alt import parse_locator
    locs = _locators(undescribed)
    assert len(locs) == 2, f"expected two undescribed images, got {locs}"
    assert [parse_locator(loc) for loc in locs] == [(DRAWING, "Image 1"), (DRAWING, "Image 2")], locs


# ── 2. approval → write → re-scan → credit, through the real path ─────────────

@pytest.fixture()
def applied(store, monkeypatch, undescribed):
    locs = _locators(undescribed)
    blob = _Blob(undescribed)
    _seed(store, {locs[0]: FIRST, locs[1]: SECOND})
    _run_lane(monkeypatch, store, blob)
    return blob, store


def test_the_saved_workbook_carries_both_approved_descriptions(applied):
    blob, _ = applied
    xml = _drawing_xml(blob.data)
    assert f'descr="{FIRST}"' in xml and f'descr="{SECOND}"' in xml
    assert 'descr="Picture"' not in xml, "a placeholder description survived alongside the real one"


def test_the_grid_survives_the_write(applied):
    """Read back through openpyxl, which had no part in writing the change. The cell values go
    through the shared-string table, so this also proves the write did not disturb it."""
    import openpyxl
    blob, _ = applied
    wb = openpyxl.load_workbook(str(_spill(blob.data)))
    assert wb.sheetnames == [SHEET], "the sheet name did not survive"
    ws = wb[SHEET]
    for ref, val in CELLS.items():
        assert ws[ref].value == val, f"{ref} lost its value"
    assert ws["B2"].value == 128 and ws["B3"].value == 96


def test_the_workbook_still_opens(applied):
    blob, _ = applied
    assert zipfile.ZipFile(_spill(blob.data)).testzip() is None


def test_a_second_real_assessment_no_longer_reports_1_1_1(applied):
    """THE claim: a fresh assessment of the SAVED bytes, not the writer's return value."""
    blob, _ = applied
    assert "1.1.1" not in _assess(blob.data)


def test_the_row_is_credited_and_the_copy_is_stored(applied):
    blob, store = applied
    assert store.count_unapplied_approved_values(SID, FILE) == 0
    assert blob.uploads


# ── 3. where the lane must NOT credit ─────────────────────────────────────────
#
# The bytes an uncredited lane wrote are DISCARDED — `_apply_one_value_kind` returns the
# original `working` and uploads nothing — so the control below establishes what the write
# produces by calling the writer directly, then runs the real lane on the same approval.

def test_a_partial_write_is_not_credited_because_the_criterion_still_fails(store, monkeypatch,
                                                                           undescribed):
    """Two undescribed images, one approved. The write succeeds and the workbook genuinely
    improves — and 1.1.1 still fails, because the other image is still silent. A lane crediting
    on the write would certify a workbook that still fails the criterion."""
    from apply_alt import apply_alt_text
    locs = _locators(undescribed)

    written, applied, _ = apply_alt_text(undescribed, {locs[0]: FIRST})
    assert applied, "the writer refused the value, so this control is not about crediting"
    assert f'descr="{FIRST}"' in _drawing_xml(written)
    assert "1.1.1" in _assess(written), (
        "the fixture no longer fails after describing one of two images — this control cannot "
        "distinguish a withheld credit from a cleared one")

    blob = _Blob(undescribed)
    _seed(store, {locs[0]: FIRST})
    _run_lane(monkeypatch, store, blob)

    assert store.count_unapplied_approved_values(SID, FILE) == 1
    assert store.mark_file_compliant_if_reviewed(SID, FILE) is False
    assert not blob.uploads
    assert blob.data == undescribed


def test_an_already_described_workbook_is_left_byte_identical(store, monkeypatch):
    clean = _book(pictures=2, alts=(FIRST, SECOND))
    assert _locators(clean) == []
    blob = _Blob(clean)
    _seed(store, {})
    _run_lane(monkeypatch, store, blob)
    assert blob.data == clean and not blob.uploads


# ── 4. a broken engine earns nothing ──────────────────────────────────────────

@pytest.mark.parametrize("name,script,timeout", [
    ("cannot be launched", None, None),
    ("exits non-zero", "#!/bin/sh\necho boom >&2\nexit 9\n", None),
    ("hangs past the timeout", "#!/bin/sh\nsleep 30\n", "2"),
])
def test_a_broken_office_analyser_never_credits_this_lane_either(monkeypatch, undescribed,
                                                                 name, script, timeout):
    """Re-asserted per lane rather than assumed to inherit: the fail-open #1058 closed lived in
    ONE shared seam, so a regression there takes every lane at once."""
    import stat as _stat

    import scanner
    if script is None:
        monkeypatch.setattr(scanner, "DOTNET", "/nonexistent/dotnet", raising=False)
    else:
        fake = Path(tempfile.mkdtemp()) / "dotnet"
        fake.write_text(script)
        fake.chmod(fake.stat().st_mode | _stat.S_IEXEC)
        monkeypatch.setattr(scanner, "DOTNET", str(fake), raising=False)
    if timeout:
        monkeypatch.setenv("ACP_OFFICE_CLI_TIMEOUT", timeout)

    from proposals import verify_residual_scs
    residual = verify_residual_scs(undescribed, FILE)
    assert residual is not None, (
        f"an office CLI that {name} made the re-scan return None — every approved value on this "
        f"lane would be credited on a scan that never happened")
    assert "1.1.1" in residual
