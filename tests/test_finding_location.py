"""Every finding should be able to answer "where is it?" — and say nothing when it cannot.

WHY THIS EXISTS. A census over the ground-truth corpora (xlsx, pptx, pdf; gen_sc_corpus has a
different builder API and was not included) found 4 of 48 findings carrying a `location`, all four
of them from the vendored PDF analyser. Every first-party detector emitted none — including ones
holding the answer already: `pptx_checks` iterates ppt/slides/slideN.xml and knows the slide, and
an xlsx `<hyperlink>` tag carries the very cell reference the writer uses to fix it.

Nothing downstream needed building. `store._loc()` already persists `location` or `locator`, and
frontend/src/EvidenceCard.jsx already renders `card.location` as a 📍 chip — with a comment
stating the rule this file enforces: "Rendered only when the analysers attributed a page — the
reviewer gets no location rather than a wrong one." So the gap was entirely detector-side, and
the tests below are about attaching what was already known and refusing to invent the rest.

WHAT A LOCATION IS FOR. It answers the first thing a reviewer asks and the one thing a finding
cannot be acted on without: which slide to open, which cell to look at. A finding that names a
count and an example without saying where either lives is a report, not a task.
"""
from __future__ import annotations

import io
import sys
import tempfile
import zipfile
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "api"))

import office_structure as osx  # noqa: E402

pytest.importorskip("openpyxl")
pytest.importorskip("docx")


def _find(findings: list[dict], rule_id: str) -> dict:
    hit = [f for f in findings if f.get("ruleId") == rule_id]
    assert hit, f"{rule_id} did not fire; got {[f.get('ruleId') for f in findings]}"
    return hit[0]


# ── the shared formatter ─────────────────────────────────────────────────────

def test_no_known_place_yields_no_location_at_all():
    """The load-bearing case. `None` is what makes the card render no chip; "" or "unknown"
    would put an empty or lying 📍 in front of a reviewer."""
    assert osx._where([]) is None
    assert osx._where([None, None]) is None
    assert osx._where(["", None]) is None


def test_one_place_is_named_and_several_are_counted():
    assert osx._where(["Slide 3"]) == "Slide 3"
    assert osx._where(["Slide 3", "Slide 7"]) == "Slide 3 (+1 more)"


def test_repeats_of_one_place_collapse():
    """Three bad links on one slide is one place, not "Slide 3 (+2 more)" — which would tell a
    reviewer to go looking for slides that do not exist."""
    assert osx._where(["Slide 3", "Slide 3", "Slide 3"]) == "Slide 3"


# ── pptx ─────────────────────────────────────────────────────────────────────

def _deck(tmp: Path) -> Path:
    def slide(body: str) -> str:
        return ('<p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
                'xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main" '
                'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
                f'<p:cSld><p:spTree>{body}</p:spTree></p:cSld></p:sld>')

    titled = ('<p:sp><p:nvSpPr><p:nvPr><p:ph type="title"/></p:nvPr></p:nvSpPr>'
              '<p:txBody><a:p><a:r><a:t>Real title</a:t></a:r></a:p></p:txBody></p:sp>')
    empty = ('<p:sp><p:nvSpPr><p:nvPr><p:ph type="title"/></p:nvPr></p:nvSpPr>'
             '<p:txBody><a:p><a:r><a:t></a:t></a:r></a:p></p:txBody></p:sp>')
    def link(text: str) -> str:
        return ('<p:sp><p:txBody><a:p><a:r><a:rPr><a:hlinkClick r:id="rId1"/></a:rPr>'
                f'<a:t>{text}</a:t></a:r></a:p></p:txBody></p:sp>')

    rels = ('<?xml version="1.0"?><Relationships '
            'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
            '2006/relationships/hyperlink" Target="https://example.org/x" '
            'TargetMode="External"/></Relationships>')

    p = tmp / "deck.pptx"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        # A DESCRIPTIVE link on slide 1, earlier than the vague one. Without it, "the place of
        # the first vague link" and "the place of the first link" are the same slide and a
        # writer that confused the two would pass — which is exactly what happened: the first
        # version of this fixture had one link, and the bite check for that confusion did not
        # bite. A control that cannot distinguish the two answers is not a control.
        z.writestr("ppt/slides/slide1.xml", slide(titled + link("Q3 accessibility report")))
        z.writestr("ppt/slides/slide2.xml", slide(empty))            # 2.4.6 here
        z.writestr("ppt/slides/slide3.xml", slide(titled + link("click here")))   # 2.4.4 here
        z.writestr("ppt/slides/_rels/slide1.xml.rels", rels)
        z.writestr("ppt/slides/_rels/slide3.xml.rels", rels)
    p.write_bytes(buf.getvalue())
    return p


def test_an_empty_slide_title_names_its_slide(tmp_path):
    f = _find(osx.pptx_checks(_deck(tmp_path)), "PPTX_TITLE_EMPTY")
    assert f["location"] == "Slide 2"


def test_a_vague_link_names_the_slide_it_is_on_not_the_first(tmp_path):
    """The link is on slide 3 and the deck has three slides, so a writer that reported the first
    slide, the last slide, or a constant would still look plausible against a one-slide deck."""
    f = _find(osx.pptx_checks(_deck(tmp_path)), "PPTX_LINK_PURPOSE_VAGUE")
    assert f["location"] == "Slide 3"


# ── xlsx ─────────────────────────────────────────────────────────────────────

def test_a_vague_cell_link_names_the_sheet_and_the_cell(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Q3 Findings"
    ws["B2"] = "click here"
    ws["B2"].hyperlink = "https://example.org/report"
    p = tmp_path / "book.xlsx"
    wb.save(p)

    f = _find(osx.xlsx_structure_checks(p), "XLSX_LINK_PURPOSE_VAGUE")
    assert f["location"] == "Sheet “Q3 Findings” cell B2"


def test_the_sheet_name_is_resolved_through_the_rels_not_guessed_from_the_part_number(tmp_path):
    """THE case that separates a real mapping from an ordinal guess.

    `<sheet>` elements carry an r:id, and the tab order need not match the part names — a
    workbook whose tabs were reordered or deleted keeps the original sheetN.xml names. Here the
    FIRST tab ("Summary") is xl/worksheets/sheet2.xml and the second ("Detail") is sheet1.xml.
    The vague link lives in sheet1.xml, so the honest answer is "Detail"; reading the first
    <sheet> element's name off the part number gives "Summary" and sends the reviewer to the
    wrong tab.
    """
    wb = ('<?xml version="1.0"?><workbook '
          'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
          'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets>'
          '<sheet name="Summary" sheetId="1" r:id="rId1"/>'
          '<sheet name="Detail" sheetId="2" r:id="rId2"/>'
          '</sheets></workbook>')
    wb_rels = ('<?xml version="1.0"?><Relationships '
               'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
               '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
               '2006/relationships/worksheet" Target="worksheets/sheet2.xml"/>'
               '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/'
               '2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
               '</Relationships>')
    sheet = ('<?xml version="1.0"?><worksheet '
             'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
             'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
             '<sheetData><row r="2"><c r="B2" t="inlineStr"><is><t>click here</t></is></c></row>'
             '</sheetData>'
             '<hyperlinks><hyperlink ref="B2" r:id="rId1"/></hyperlinks></worksheet>')
    rels = ('<?xml version="1.0"?><Relationships '
            'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
            '2006/relationships/hyperlink" Target="https://example.org/x" '
            'TargetMode="External"/></Relationships>')

    p = tmp_path / "reordered.xlsx"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as z:
        z.writestr("xl/workbook.xml", wb)
        z.writestr("xl/_rels/workbook.xml.rels", wb_rels)
        z.writestr("xl/worksheets/sheet1.xml", sheet)
        z.writestr("xl/worksheets/sheet2.xml",
                   '<?xml version="1.0"?><worksheet><sheetData/></worksheet>')
        z.writestr("xl/worksheets/_rels/sheet1.xml.rels", rels)
    p.write_bytes(buf.getvalue())

    f = _find(osx.xlsx_structure_checks(p), "XLSX_LINK_PURPOSE_VAGUE")
    assert f["location"] == "Sheet “Detail” cell B2", (
        "the sheet name was taken from the part number rather than resolved through the "
        "workbook relationships")


def test_default_sheet_tabs_name_the_tabs_they_are_about(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    wb.active.title = "Sheet1"
    wb.create_sheet("Sheet2")
    wb.create_sheet("Sheet3")
    p = tmp_path / "default.xlsx"
    wb.save(p)

    f = _find(osx.xlsx_structure_checks(p), "XLSX_DEFAULT_LABELS")
    assert f["location"] == "Sheet “Sheet1” (+2 more)"


def test_a_workbook_with_no_vague_links_reports_nothing_to_locate(tmp_path):
    """The control: the location must not be the thing that makes a finding appear."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Q3 Findings"
    ws["B2"] = "Q3 accessibility report"
    ws["B2"].hyperlink = "https://example.org/report"
    p = tmp_path / "clean.xlsx"
    wb.save(p)
    assert [f for f in osx.xlsx_structure_checks(p)
            if f.get("ruleId") == "XLSX_LINK_PURPOSE_VAGUE"] == []


# ── the deliberate absence ───────────────────────────────────────────────────

def test_docx_link_findings_still_carry_no_location_and_that_is_correct(tmp_path):
    """docx is deliberately NOT in this change. `_docx_hyperlinks` walks the body, headers,
    footers and notes and does not track which paragraph a link came from, so there is no
    position to attach — and the rule is that a detector which cannot say where something is
    says nothing rather than something plausible.

    Asserted rather than left implicit, because the natural next commit is "finish the job by
    adding docx", and doing that by reporting the PART ("word/document.xml") would technically
    populate the field while telling a reviewer nothing they can act on. When docx gains a real
    paragraph reference this test fails, and that is the moment to delete it.
    """
    from docx import Document
    from docx.oxml import OxmlElement
    from docx.oxml.ns import qn

    p = tmp_path / "doc.docx"
    doc = Document()
    para = doc.add_paragraph()
    rid = para.part.relate_to("https://example.org/x",
                             "http://schemas.openxmlformats.org/officeDocument/2006/"
                             "relationships/hyperlink", is_external=True)
    link = OxmlElement("w:hyperlink")
    link.set(qn("r:id"), rid)
    run = OxmlElement("w:r")
    t = OxmlElement("w:t")
    t.text = "click here"
    run.append(t)
    link.append(run)
    para._p.append(link)
    doc.save(str(p))

    f = _find(osx.checks_for(p, ".docx"), "DOCX_LINK_PURPOSE_VAGUE")
    assert "location" not in f, (
        "docx now reports a location — if it is a real paragraph reference, delete this test; "
        "if it is the part name, that is not a location a reviewer can use")
