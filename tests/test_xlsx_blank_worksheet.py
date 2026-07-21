"""Regression guard for BlankWorksheetRule (XLSX-BLANK-001, WCAG 1.3.2).

Detection lives in the vendored .NET analyser
(engine/office-analysers/DigitalA11y.Analysers.DotNet/Xlsx/Rules/BlankWorksheetRule.cs).
The rule must check actual cell content — not openpyxl/Excel's cached dimension
("used range") record, which stays inflated after cells are cleared — must not flag a
sheet that holds only a chart/image with no cell data, and must not flag a hidden sheet
(found via the test-corpus oracle fixture xlsx-noncompliant.xlsx, which has a genuinely
blank sheet named "Hidden" — no user ever tabs to it, so it must stay silent).

This builds a workbook with sheets covering each of those cases and asserts the rule
fires for exactly the genuinely blank, visible ones. It needs the built .NET CLI to run,
so it self-skips when the analysers didn't run (CLI not built).
"""
import io
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "api"))
import scanner  # noqa: E402


def _rule_ids(fdict) -> list[str]:
    return [i.get("ruleId") for i in fdict.get("issues", [])]


@pytest.fixture(scope="module")
def scanned(tmp_path_factory):
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image

    d = tmp_path_factory.mktemp("xlsx-blank")
    wb = Workbook()

    data_ws = wb.active
    data_ws.title = "Data"
    for row in [["Region", "Total"], ["North", 10], ["South", 20]]:
        data_ws.append(row)

    blank_ws = wb.create_sheet("Blank")  # every cell empty, no drawings — must be flagged

    chart_ws = wb.create_sheet("ChartOnly")  # no cell data of its own, but anchors a chart
    chart = BarChart()
    chart.add_data(Reference(data_ws, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    chart_ws.add_chart(chart, "A1")

    picture_ws = wb.create_sheet("PictureOnly")  # no cell data, but anchors an image
    buf = io.BytesIO()
    Image.new("RGB", (100, 60), (10, 10, 10)).save(buf, format="PNG")
    picture_ws.add_image(XLImage(io.BytesIO(buf.getvalue())), "A1")

    stale_range_ws = wb.create_sheet("StaleRange")  # every cell cleared but max_row/col stays inflated
    stale_range_ws["C5"] = "temp"
    stale_range_ws["C5"] = None

    hidden_blank_ws = wb.create_sheet("HiddenBlank")  # blank AND hidden — no tab to confuse a user
    hidden_blank_ws.sheet_state = "hidden"

    name = "blank-worksheet.xlsx"
    wb.save(str(d / name))

    fdict, _ = scanner.analyse_and_assess(d, name, detect_pii=False)
    return {"rule_ids": _rule_ids(fdict),
            "locations": [i.get("location") for i in fdict.get("issues", [])
                          if i.get("ruleId") == "XLSX-BLANK-001"],
            "engine_ran": any(str(i.get("wcag", "")).startswith("SC_") for i in fdict.get("issues", []))}


def test_genuinely_blank_sheet_is_flagged(scanned):
    if not scanned["engine_ran"]:
        pytest.skip("the .NET office analysers did not run (CLI not built)")
    assert scanned["rule_ids"].count("XLSX-BLANK-001") >= 1, \
        "a genuinely empty sheet was not flagged (BlankWorksheetRule regressed)"


def test_data_sheet_is_not_flagged(scanned):
    if not scanned["engine_ran"]:
        pytest.skip("the .NET office analysers did not run (CLI not built)")
    assert not any("Data" in str(loc) for loc in scanned["locations"]), \
        "the sheet holding real data was incorrectly flagged as blank"


def test_chart_only_sheet_is_not_flagged(scanned):
    if not scanned["engine_ran"]:
        pytest.skip("the .NET office analysers did not run (CLI not built)")
    assert not any("ChartOnly" in str(loc) for loc in scanned["locations"]), \
        "a sheet with only a chart (no cell data) was incorrectly flagged as blank"


def test_picture_only_sheet_is_not_flagged(scanned):
    if not scanned["engine_ran"]:
        pytest.skip("the .NET office analysers did not run (CLI not built)")
    assert not any("PictureOnly" in str(loc) for loc in scanned["locations"]), \
        "a sheet with only an image (no cell data) was incorrectly flagged as blank"


def test_stale_used_range_is_flagged(scanned):
    """A sheet whose only cell was cleared keeps an inflated dimension/used-range record.
    The rule must check live cell content, not that cached record, so this must still fire."""
    if not scanned["engine_ran"]:
        pytest.skip("the .NET office analysers did not run (CLI not built)")
    assert any("StaleRange" in str(loc) for loc in scanned["locations"]), \
        "a sheet with only cleared cells was not flagged — rule may be trusting the stale used-range record"


def test_hidden_blank_sheet_is_not_flagged(scanned):
    """No user ever tabs to a hidden sheet, blank or not — flagging it is a false alarm
    (this is what the xlsx-noncompliant.xlsx oracle fixture's "Hidden" sheet caught)."""
    if not scanned["engine_ran"]:
        pytest.skip("the .NET office analysers did not run (CLI not built)")
    assert not any("HiddenBlank" in str(loc) for loc in scanned["locations"]), \
        "a hidden blank sheet was incorrectly flagged — hidden sheets have no navigation exposure"


def test_exactly_the_two_visible_blank_sheets_are_flagged(scanned):
    if not scanned["engine_ran"]:
        pytest.skip("the .NET office analysers did not run (CLI not built)")
    assert scanned["rule_ids"].count("XLSX-BLANK-001") == 2, \
        f"expected exactly Blank + StaleRange to fire, got: {scanned['rule_ids']}"
