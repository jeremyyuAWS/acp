#!/usr/bin/env python3
"""A LABELLED .xlsx corpus — the second format to get ground truth, after .docx.

WHY THIS EXISTS. `scripts/gen_fixture_coverage.py` reports coverage per (criterion, format) pair.
When this corpus was written .docx was complete and xlsx / pptx / pdf were at zero; this was the
start of the xlsx half, and it now declares ELEVEN of fifteen.

IT IS DELIBERATELY PARTIAL, AND THE LIMIT IS VERIFICATION, NOT EFFORT. Every fixture seeds a
violation a detector is confirmed to fire on, with an adversarial counterpart confirmed to leave
it silent. A fixture whose seeded violation nobody can confirm is caught would inflate the very
coverage number this corpus exists to report honestly.

    1.1.1   an image with no descr                       office_non_text_content_checks
    1.4.1   colour-scale conditional formatting          office_color_only_checks
    1.4.3   text/fill contrast under 4.5:1               xlsx_contrast_checks
    1.4.11  a faint shape outline on its fill            xlsx_nontext_contrast_checks
    2.1.2   an embedded form control                     office_control_review_checks
    2.4.4   vague or raw-URL hyperlink labels            xlsx_structure_checks
    2.4.6   default SheetN tabs / ColumnN headers        xlsx_structure_checks
    1.3.3   a sensory-only instruction, in a cell        textchecks.detect_sensory
    4.1.2   an embedded form control (same fixture)      office_control_review_checks
    ----    confirmed only where the .NET analyser is built (DECLARED_ENGINE) ----
    2.4.2   no dc:title in the core properties           XLSX-TITLE-001
    3.1.1   no dc:language in the core properties        XLSX-LANG-001

WHERE THE CONFIRMATION HAPPENS is the one asymmetry, and it is recorded rather than smoothed
over. The first eight are proven wherever the suite runs. The last two have NO first-party Python
detector on any Office format, so they are proven by the .NET analyser and skipped on a bare
checkout — see DECLARED_ENGINE. They earn that exception by being certifying pairs: 2.4.2 and
3.1.1 are two of the seventeen (criterion, format) pairs in the whole preset that can return a
PASS, so before these fixtures a clean scan certified a spreadsheet against criteria nothing in
the suite checked.

The four still missing need something this box does not have: 1.3.1 and 1.3.2 the .NET
analyser, 1.4.5 tesseract, 3.1.2 langdetect. 1.3.3 was in that list until it was checked —
it is a TEXT predicate (textchecks.detect_sensory) with no engine behind it at all, and it is
declared above. Reachability is worth testing rather than inferring from a neighbour.

THE VOCABULARY IS NOT PASS/FAIL, and assuming it is would invalidate the labels. ACP answers four
things, and which are reachable is a property of the pair. On .xlsx only FIVE of the fifteen
Core-17 pairs can ever return PASS:

    1.3.1  1.3.2  1.4.3  2.4.2  3.1.1

Ten are REVIEW-lane: a review detector does not certify conformance (ADR 0016), so a CLEAN file
there resolves to REVIEW and NEVER to PASS. Every declaration below is checked against
`corpus_expectations.possible_verdicts()` at generation time — a fixture that expects PASS on
1.4.1 fails to build. Without that check the corpus would report a false failure on every run
forever, and it would read as a product defect rather than a manifest defect.

Note what that means for 1.4.1, 2.4.4 and 2.4.6 here: their violation fixtures expect FAIL or
REVIEW per their lane, and their adversarial counterparts expect REVIEW — not PASS — because the
engine is not permitted to certify those pairs at all.

Run:
    python scripts/gen_xlsx_corpus.py --out ~/Downloads/acp-xlsx-eval/sc-corpus
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "api"))
sys.path.insert(0, str(ROOT / "scripts"))

import corpus_expectations as ce  # noqa: E402

FMT = "xlsx"

# A readable sheet on a white ground — dark slate on white is ~12.6:1, so nothing a fixture does
# not deliberately break trips the contrast check.
INK = "FF1A1F26"
PAPER = "FFFFFFFF"


DOC_TITLE = "Q3 Benefits Summary"
DOC_LANG = "en-GB"          # PackageProperties.Language — what XLSX-LANG-001 actually reads


def _wb(title: str | None = DOC_TITLE, language: str | None = DOC_LANG):
    """A workbook clean on everything the corpus is not deliberately breaking: a named sheet (so
    2.4.6's default-tab rule stays quiet), a document title, a declared language, and legible text.

    THE LANGUAGE IS NOT DECORATION, and its absence was a real defect in this corpus. openpyxl
    leaves `properties.language` at None, and XLSX-LANG-001 reports 3.1.1 for exactly that — so
    before this every fixture here ALSO carried a 3.1.1 finding, and the corpus's single-criterion
    labels were true only because the .NET engine is absent from a bare container. In CI, where
    the engine IS built, each one was mislabelled: a fixture declared as "1.4.3 FAIL and nothing
    else" produced two findings, which is precisely the error a ground-truth corpus exists to make
    impossible. Nothing caught it because the first-party tests read `checks_for`, and the engine
    reports through a different path.
    """
    from openpyxl import Workbook
    wb = Workbook()
    wb.properties.title = title
    wb.properties.language = language
    ws = wb.active
    ws.title = "Summary"
    return wb, ws


def _say(ws, ref: str, text: str, *, colour: str = INK, fill: str = PAPER):
    from openpyxl.styles import Font, PatternFill
    ws[ref] = text
    ws[ref].font = Font(color=colour)
    ws[ref].fill = PatternFill("solid", fgColor=fill)


# A 1x1 PNG, inline. Keeps a binary asset out of the repo — the fixture needs an image to embed,
# not a picture of anything, and a file whose provenance a reviewer has to take on trust is
# exactly what a generated corpus avoids.
_PNG_1PX = bytes.fromhex(
    "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c489"
    "0000000d49444154789c6360f8cf0000010101001b0b8b5c0000000049454e44ae426082")


def _inject(path: Path, parts: dict[str, str]) -> None:
    """Rewrite the saved workbook with extra zip parts.

    openpyxl cannot author a DrawingML shape or a form-control part, and both are what their
    detectors read. Post-processing the zip is the honest way to reach them — the .docx generator
    already does the same thing for its own unreachable case (`_strip_title`). The alternative is
    leaving those criteria uncovered because a library lacks a setter, which is a fact about
    openpyxl rather than about ACP.
    """
    import shutil
    import tempfile
    import zipfile
    tmp = Path(tempfile.mkdtemp()) / path.name
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            zout.writestr(item, zin.read(item.filename))
        for name, body in parts.items():
            zout.writestr(name, body)
    shutil.move(str(tmp), str(path))


def _shape_drawing(outline: str, fill: str) -> str:
    """One DrawingML shape: a solid `outline` on a solid `fill`. xlsx_nontext_contrast_checks
    reports the WORST outline-on-fill ratio under 3:1 across the drawing parts."""
    return (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
        ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
        "<xdr:twoCellAnchor><xdr:sp><xdr:spPr>"
        f'<a:solidFill><a:srgbClr val="{fill}"/></a:solidFill>'
        f'<a:ln w="12700"><a:solidFill><a:srgbClr val="{outline}"/></a:solidFill></a:ln>'
        "</xdr:spPr></xdr:sp></xdr:twoCellAnchor></xdr:wsDr>")


_CTRL_PROPS = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    '<formControlPr xmlns="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"'
    ' objectType="Drop" dropStyle="combo" dx="16" fmlaLink="$B$1" sel="0" val="0"/>')


# ── 1.4.3 Contrast (Minimum) ─────────────────────────────────────────────────────

def f_contrast_fail(wb, ws):
    # #DDDDDD on white is ~1.6:1 — far under 4.5, and unambiguous rather than a boundary case.
    _say(ws, "A1", "Quarterly revenue summary for the finance committee", colour="FFDDDDDD")
    return {"1.4.3": "FAIL"}, "grey #DDDDDD on white ≈ 1.6:1, well under the 4.5:1 minimum"


def f_contrast_ok(wb, ws):
    _say(ws, "A1", "Quarterly revenue summary for the finance committee")
    # PASS is reachable here — 1.4.3 is one of the five .xlsx pairs that can certify.
    return {"1.4.3": "PASS"}, "dark slate on white ≈ 12.6:1 — comfortably over the minimum"


# ── 2.4.4 Link Purpose (In Context) ──────────────────────────────────────────────

def f_link_vague(wb, ws):
    ws.title = "Policies"
    _say(ws, "A1", "Travel policy")
    _say(ws, "A2", "click here")
    ws["A2"].hyperlink = "https://example.com/fy26-travel-policy.pdf"
    return {"2.4.4": "REVIEW"}, "a 'click here' label — the destination is unknowable from the text"


def f_link_descriptive_ok(wb, ws):
    ws.title = "Policies"
    _say(ws, "A1", "Travel policy")
    _say(ws, "A2", "Read the FY26 travel policy")
    ws["A2"].hyperlink = "https://example.com/fy26-travel-policy.pdf"
    # REVIEW, not PASS: 2.4.4 on .xlsx is a review-lane pair and cannot certify however clean the
    # file is. A fixture expecting PASS here would fail _validate, which is the point of it.
    return {"2.4.4": "REVIEW"}, "a descriptive label — must NOT be flagged (adversarial)"


# ── 2.4.6 Headings and Labels ────────────────────────────────────────────────────

def f_default_sheet_tabs(wb, ws):
    ws.title = "Sheet1"
    wb.create_sheet("Sheet2")
    wb.create_sheet("Sheet3")
    _say(ws, "A1", "Regional totals")
    return {"2.4.6": "REVIEW"}, "three default SheetN tabs — the structure labels say nothing"


def f_one_default_tab_ok(wb, ws):
    # THE EDGE CASE, and the reason the detector requires two: a lone 'Sheet1' is what Excel
    # gives every new workbook, and flagging it would fire on almost every file in an estate.
    ws.title = "Sheet1"
    _say(ws, "A1", "Regional totals")
    return {"2.4.6": "REVIEW"}, "a LONE default tab — normal, must not be flagged (adversarial)"


def f_named_tabs_ok(wb, ws):
    ws.title = "Regional totals"
    wb.create_sheet("Headcount")
    wb.create_sheet("Assumptions")
    _say(ws, "A1", "Regional totals")
    return {"2.4.6": "REVIEW"}, "three named tabs — must not be flagged (adversarial)"


# ── 1.4.1 Use of Color ───────────────────────────────────────────────────────────

def f_colour_scale(wb, ws):
    from openpyxl.formatting.rule import ColorScaleRule
    ws.title = "Risk"
    _say(ws, "A1", "Region")
    _say(ws, "B1", "Score")
    for i, v in enumerate((10, 55, 90), start=2):
        _say(ws, f"A{i}", f"Region {i - 1}")
        ws[f"B{i}"] = v
    ws.conditional_formatting.add(
        "B2:B4", ColorScaleRule(start_type="min", start_color="FFF8696B",
                                end_type="max", end_color="FF63BE7B"))
    return {"1.4.1": "REVIEW"}, "a red→green colorScale — status carried by colour alone"


def f_icon_set_ok(wb, ws):
    from openpyxl.formatting.rule import IconSetRule
    ws.title = "Risk"
    _say(ws, "A1", "Region")
    _say(ws, "B1", "Score")
    for i, v in enumerate((10, 55, 90), start=2):
        _say(ws, f"A{i}", f"Region {i - 1}")
        ws[f"B{i}"] = v
    # An iconSet pairs colour WITH a shape, so it is NOT colour-only — the detector says so
    # explicitly, and this fixture is what holds it to that.
    ws.conditional_formatting.add("B2:B4", IconSetRule("3TrafficLights1", "percent", [0, 33, 67]))
    return {"1.4.1": "REVIEW"}, "an iconSet — colour PLUS a shape, must not be flagged (adversarial)"


# ── 1.1.1 Non-text Content ───────────────────────────────────────────────────────

def f_image_no_alt(wb, ws):
    import tempfile
    from openpyxl.drawing.image import Image as XLImage
    ws.title = "Chart"
    _say(ws, "A1", "Revenue by region")
    png = Path(tempfile.mkdtemp()) / "px.png"
    png.write_bytes(_PNG_1PX)
    ws.add_image(XLImage(str(png)), "C3")
    return {"1.1.1": "REVIEW"}, "an embedded image with no descr — nothing for a screen reader"


def f_no_image_ok(wb, ws):
    ws.title = "Chart"
    _say(ws, "A1", "Revenue by region")
    _say(ws, "A2", "North 412  ·  South 388  ·  East 501")
    # A sheet with no non-text content at all must not produce a 1.1.1 finding. The obvious
    # adversarial case, and the one a "flag every sheet" regression would trip.
    return {"1.1.1": "REVIEW"}, "no images at all — must not be flagged (adversarial)"


# ── 1.4.11 Non-text Contrast ─────────────────────────────────────────────────────

def f_shape_faint_outline(wb, ws):
    ws.title = "Diagram"
    _say(ws, "A1", "Process overview")
    # #C8C8C8 on #FFFFFF ≈ 1.8:1, under the 3:1 a meaningful boundary needs.
    return ({"1.4.11": "REVIEW"}, "a shape outline #C8C8C8 on white ≈ 1.8:1, under 3:1",
            {"xl/drawings/drawing1.xml": _shape_drawing("C8C8C8", "FFFFFF")})


def f_shape_strong_outline_ok(wb, ws):
    ws.title = "Diagram"
    _say(ws, "A1", "Process overview")
    # #1A1F26 on white ≈ 15:1 — the same shape, drawn visibly. Holds the detector to reporting
    # only outlines that are actually too faint, rather than every shape it can find.
    return ({"1.4.11": "REVIEW"}, "the same shape at ≈15:1 — must not be flagged (adversarial)",
            {"xl/drawings/drawing1.xml": _shape_drawing("1A1F26", "FFFFFF")})


# ── 4.1.2 Name, Role, Value / 2.1.2 No Keyboard Trap ─────────────────────────────

def f_form_control(wb, ws):
    ws.title = "Form"
    _say(ws, "A1", "Department")
    # ONE fixture, TWO criteria: an embedded control is evidence for both the accessible-name
    # question (4.1.2) and the keyboard-trap question (2.1.2), and the detector reports both.
    # Neither can be settled from the file alone, which is why both are review-lane.
    return ({"4.1.2": "REVIEW", "2.1.2": "REVIEW"},
            "an xlsx form control — name/role and keyboard behaviour both need a human",
            {"xl/ctrlProps/ctrlProp1.xml": _CTRL_PROPS})


def f_no_controls_ok(wb, ws):
    ws.title = "Form"
    _say(ws, "A1", "Department")
    _say(ws, "A2", "Finance")
    return ({"4.1.2": "REVIEW", "2.1.2": "REVIEW"},
            "a static sheet with no controls — must not be flagged (adversarial)")


# ── 2.4.2 Page Titled and 3.1.1 Language of Page — the ENGINE-VERIFIED pairs ─────
# These two differ from everything above: no first-party Python detector exists for either on
# any Office format, so they are confirmed by the .NET analyser (XLSX-TITLE-001, XLSX-LANG-001)
# and only where it is built. That makes them the first pairs in this corpus whose label is
# proven in CI rather than on any machine — see DECLARED_ENGINE, which keeps them countable
# without letting them dilute what DECLARED means.
#
# Both are worth the asymmetry: 2.4.2 and 3.1.1 are two of the seventeen (criterion, format)
# pairs in the whole preset that can return a PASS, so a false clean result here is a
# certification, not an advisory. They were among the eight such pairs with no ground truth at all.
#
# BUILT FROM THE VENDORED RULE SOURCE, NOT THE CATALOG. config/rule-catalog.json describes
# XLSX-LANG-001 as "the workbook must declare a language in its styles.xml"; the rule itself
# (engine/office-analysers/.../Xlsx/Rules/DocumentLanguageRule.cs) reads
# `document.PackageProperties.Language` — the core-property dc:language, nowhere near styles.xml.
# A fixture written from the description would have injected a styles.xml part and detected
# nothing, which is the failure this corpus exists to prevent, arriving through the catalog.

def f_no_document_title(wb, ws):
    wb.properties.title = None
    _say(ws, "A1", "Quarterly revenue summary for the finance committee")
    return {"2.4.2": "FAIL"}, "no dc:title in the core properties — nothing identifies the file"


def f_document_title_ok(wb, ws):
    _say(ws, "A1", "Quarterly revenue summary for the finance committee")
    return {"2.4.2": "PASS"}, f"dc:title set to {DOC_TITLE!r} (adversarial)"


def f_no_document_language(wb, ws):
    wb.properties.language = None
    _say(ws, "A1", "Quarterly revenue summary for the finance committee")
    return {"3.1.1": "FAIL"}, "no language in the core properties — no pronunciation rules"


def f_document_language_ok(wb, ws):
    _say(ws, "A1", "Quarterly revenue summary for the finance committee")
    return {"3.1.1": "PASS"}, f"language set to {DOC_LANG!r} (adversarial)"


# ── 1.3.3 Sensory Characteristics — decided by the prose, not the workbook ──────
# The one criterion in this corpus that does not read the file's structure at all:
# textchecks.detect_sensory reads the EXTRACTED TEXT for an instruction that identifies a control
# only by shape, colour or position. The words are the fixture; the spreadsheet is the container.
# Deliberately identical wording to the .pptx and .pdf corpora, so a change in detector behaviour
# shows up as the same result in three places rather than three arguments about three sentences.
SENSORY_BAD = ("To continue, click the round green button on the right. "
               "See the box below for the payment terms.")
SENSORY_OK = ("To continue, choose Submit under Payment options. "
              "The payment terms are in the Payment terms section.")


def f_sensory_instruction(wb, ws):
    ws.title = "Payments"
    _say(ws, "A1", "Payment instructions")
    _say(ws, "A2", SENSORY_BAD)
    return ({"1.3.3": "FAIL"},
            "an instruction identifying a control only by shape, colour and position")


def f_sensory_instruction_ok(wb, ws):
    ws.title = "Payments"
    _say(ws, "A1", "Payment instructions")
    _say(ws, "A2", SENSORY_OK)
    return ({"1.3.3": "REVIEW"},
            "the same instruction naming the control and the section (adversarial)")



# ── 1.4.5 Images of Text — decided by OCR, not by the container ─────────────────
# Like 1.3.3, this criterion reads the DOCUMENT'S PIXELS rather than its structure: ocr.py runs
# tesseract over each embedded raster and flags any carrying >= ocr._MIN_WORDS (10) real words.
# The image is the fixture and the sheet is the container, so the wording is identical across
# the .xlsx, .pptx and .pdf corpora — see tests/test_images_of_text_corpus.py, which asserts that
# sameness so one detector change reads as one result in three places.
#
# THE ALT IS CORRECT ON PURPOSE, and that is what keeps the fixture single-criterion. 1.4.5 is
# alt-agnostic — `images_of_text` never looks at a descr — so an image of prose fails 1.4.5
# whether or not it is described. Leaving the alt off would fail 1.1.1 as well and the fixture
# would measure two things at once. (The .docx corpus's equivalent DOES declare both, which is
# correct there: gen_sc_corpus is scored by score_assessment rather than by the single-criterion
# sweep these three corpora use.)
#
# THE WORD FLOOR IS THE TRAP. _MIN_WORDS counts OCR'd TOKENS, not what was drawn — dates and
# phone numbers are not words. The .docx corpus hit this: three short lines recovered six words,
# so 1.4.9 (floor 3) fired and 1.4.5 (floor 10) did not, and the fixture read as an engine bug.
# Hence prose, and enough of it: this image OCRs to 29 words, with room for the odd merge
# ("begins on" comes back as "beginson").
IMAGE_OF_TEXT_PROSE = ("Open enrollment for the coming plan year",
                       "begins on the first of March and closes",
                       "on the last day of the same month. Call",
                       "the benefits office to change your plan.")
IMAGE_OF_TEXT_ALT = ("Open enrollment runs from 1 March to 31 March; call the benefits office "
                     "to change plan.")
# WCAG 1.4.5 exempts logotypes, and ocr._MIN_WORDS is what encodes that exemption here: two words
# is under the floor, so the control is clean for the same reason a real logo would be.
LOGO_TEXT = ("UT Health",)


def _text_png(lines, size=(900, 360)) -> bytes:
    """A PNG with real, OCR-legible text baked into it.

    Scaled up AFTER drawing because PIL's default bitmap font is too small for tesseract to read
    reliably at native size. An unreadable image-of-text fixture silently becomes a no-text
    fixture, and 1.4.5 then stops firing for a reason that has nothing to do with the detector —
    which is the failure mode a ground-truth corpus exists to make impossible."""
    import io
    from PIL import Image, ImageDraw
    im = Image.new("RGB", size, "white")
    d = ImageDraw.Draw(im)
    y = 20
    for line in lines:
        d.text((20, y), line, fill="black")
        y += 70
    im = im.resize((size[0] * 2, size[1] * 2), Image.LANCZOS)
    buf = io.BytesIO()
    im.save(buf, format="PNG")
    return buf.getvalue()


def _text_png_file(lines, size=(900, 360)):
    import tempfile
    png = Path(tempfile.mkdtemp()) / "image-of-text.png"
    png.write_bytes(_text_png(lines, size))
    return png


def f_image_of_text(wb, ws):
    from openpyxl.drawing.image import Image as XLImage
    ws.title = "Notice"
    _say(ws, "A1", "Enrollment notice")
    img = XLImage(str(_text_png_file(IMAGE_OF_TEXT_PROSE)))
    img.desc = IMAGE_OF_TEXT_ALT
    ws.add_image(img, "C3")
    return ({"1.4.5": "FAIL"},
            "a screenshot of prose pasted into a sheet, correctly described. The alt makes it "
            "reachable but not RESIZABLE — enlarging it pixelates rather than reflows, which is "
            "what 1.4.5 is about")


def f_image_of_text_logo_ok(wb, ws):
    from openpyxl.drawing.image import Image as XLImage
    ws.title = "Notice"
    _say(ws, "A1", "Enrollment notice")
    img = XLImage(str(_text_png_file(LOGO_TEXT, size=(300, 100))))
    img.desc = "UT Health"
    ws.add_image(img, "C3")
    return ({"1.4.5": "REVIEW"},
            "a logotype with its text as the alt (adversarial). WCAG 1.4.5 exempts logos, so a "
            "finding here is a false positive")


FIXTURES = [
    ("sensory-instruction",  f_sensory_instruction,    "violation"),
    ("sensory-instruction-ok", f_sensory_instruction_ok, "adversarial"),
    ("no-document-title",    f_no_document_title,      "violation"),
    ("document-title-ok",    f_document_title_ok,      "adversarial"),
    ("no-document-language", f_no_document_language,   "violation"),
    ("document-language-ok", f_document_language_ok,   "adversarial"),
    ("contrast-fail",        f_contrast_fail,          "violation"),
    ("contrast-ok",          f_contrast_ok,            "clean"),
    ("link-vague",           f_link_vague,             "violation"),
    ("link-descriptive-ok",  f_link_descriptive_ok,    "adversarial"),
    ("sheet-tabs-default",   f_default_sheet_tabs,     "violation"),
    ("sheet-tab-single-ok",  f_one_default_tab_ok,     "edge"),
    ("sheet-tabs-named-ok",  f_named_tabs_ok,          "adversarial"),
    ("colour-scale-only",    f_colour_scale,           "violation"),
    ("colour-icon-set-ok",   f_icon_set_ok,            "adversarial"),
    ("image-no-alt",         f_image_no_alt,           "violation"),
    ("image-of-text",        f_image_of_text,          "violation"),
    ("image-of-text-logo-ok", f_image_of_text_logo_ok, "adversarial"),
    ("no-image-ok",          f_no_image_ok,            "adversarial"),
    ("shape-faint-outline",  f_shape_faint_outline,    "violation"),
    ("shape-strong-outline-ok", f_shape_strong_outline_ok, "adversarial"),
    ("form-control",         f_form_control,           "violation"),
    ("no-controls-ok",       f_no_controls_ok,         "adversarial"),
]

# The criteria this corpus declares. Kept explicit so gen_fixture_coverage and the tests agree
# with the generator about what it claims, rather than each deriving it separately.
DECLARED = ("1.1.1", "1.3.3", "1.4.1", "1.4.11", "1.4.3", "1.4.5", "2.1.2", "2.4.4",
            "2.4.6", "4.1.2")

# Declared, but confirmed only where the .NET Office analyser is built — CI, not a bare
# container. Kept in a SEPARATE tuple rather than folded into DECLARED so one number keeps one
# meaning: DECLARED is "a detector was driven against this fixture anywhere the suite runs",
# and merging the two would quietly make the coverage column mean two different things in the
# same row. gen_fixture_coverage counts both and reports the split.
DECLARED_ENGINE = ("2.4.2", "3.1.1")


def _validate(name: str, expectations: dict[str, str]) -> list[str]:
    """Reject an expectation the engine could never produce — the whole reason
    corpus_expectations exists. A manifest that expects PASS on a REVIEW-lane pair does not
    describe a product bug; it describes a manifest bug, and it would report a false failure on
    every run until somebody re-derived this by hand."""
    problems = []
    for sc, verdict in expectations.items():
        allowed = ce.possible_verdicts(sc, FMT)
        if verdict not in allowed:
            problems.append(
                f"{name}: expects {sc}={verdict} but the engine can only emit "
                f"{','.join(sorted(allowed))} for ({sc}, {FMT})")
    return problems


def build_all(docs: Path) -> tuple[list[dict], list[str]]:
    """Write every fixture and return (manifest rows, validation problems)."""
    docs.mkdir(parents=True, exist_ok=True)
    manifest, problems = [], []
    for name, build, kind in FIXTURES:
        wb, ws = _wb()
        # A builder returns (expectations, note) or, when the case needs a zip part openpyxl
        # cannot author, (expectations, note, parts). Unpacked by length rather than by a flag
        # so a builder reads the same either way.
        built = build(wb, ws)
        expectations, note = built[0], built[1]
        parts = built[2] if len(built) > 2 else None
        path = docs / f"{name}.xlsx"
        wb.save(path)
        if parts:
            _inject(path, parts)
        problems += _validate(name, expectations)
        manifest.append({"file": f"docs/{name}.xlsx", "name": name, "kind": kind,
                         "format": FMT, "expect": expectations, "note": note})
    return manifest, problems


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("--out", type=Path,
                    default=Path.home() / "Downloads" / "acp-xlsx-eval" / "sc-corpus")
    args = ap.parse_args()

    manifest, problems = build_all(args.out / "docs")
    for row in manifest:
        print(f"  {row['kind']:11} {row['name']:22} {row['expect']}")

    if problems:
        print("\nIMPOSSIBLE EXPECTATIONS — fix the fixture, not the engine:", file=sys.stderr)
        for p in problems:
            print(f"  {p}", file=sys.stderr)
        return 1

    scs = sorted(sc for sc, f in ce.pol.SCOPE_PRESETS["acp-core-17"].items() if FMT in f)
    (args.out / "manifest.json").write_text(json.dumps({
        "format": FMT, "fixtures": manifest,
        "lanes": {sc: {"clean_verdict": ce.clean_verdict(sc, FMT),
                       "possible": sorted(ce.possible_verdicts(sc, FMT)),
                       "can_pass": ce.can_ever_pass(sc, FMT)} for sc in scs},
    }, indent=1) + "\n")
    kinds = {}
    for row in manifest:
        kinds[row["kind"]] = kinds.get(row["kind"], 0) + 1
    print(f"\n{len(manifest)} fixtures: " + ", ".join(f"{n} {k}" for k, n in sorted(kinds.items())))
    print(f"declares {len(DECLARED)} of {len(scs)} applicable .xlsx pairs: {', '.join(DECLARED)}")
    print(f"wrote {args.out / 'manifest.json'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
